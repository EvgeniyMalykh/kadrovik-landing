import re
from io import BytesIO
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.employees.models import Employee, Department


# ── Column definitions ──────────────────────────────────────────────

COLUMNS = [
    # (header,          field_name,       example1,            example2,            required)
    ("Фамилия",         "last_name",      "Иванов",            "Петрова",           True),
    ("Имя",             "first_name",     "Иван",              "Мария",             True),
    ("Отчество",        "middle_name",    "Иванович",          "Сергеевна",         False),
    ("Должность",       "position",       "Менеджер",          "Бухгалтер",         True),
    ("Дата приёма",     "hire_date",      "01.03.2024",        "15.06.2023",        True),
    ("Оклад",           "salary",         "50000",             "65000",             False),
    ("Телефон",         "phone",          "+79001234567",      "+79009876543",      False),
    ("Email",           "email",          "ivan@example.com",  "maria@example.com", False),
    ("Отдел",           "department",     "Продажи",           "Бухгалтерия",       False),
    ("ИНН",             "inn",            "770123456789",      "501234567890",      False),
    ("СНИЛС",           "snils",          "123-456-789 00",    "987-654-321 00",    False),
    ("Дата рождения",   "birth_date",     "15.05.1990",        "22.11.1985",        False),
]

HEADER_NAMES = [c[0] for c in COLUMNS]
EXAMPLE_ROWS = [
    [c[2] for c in COLUMNS],
    [c[3] for c in COLUMNS],
]
REQUIRED_FIELDS = {c[0] for c in COLUMNS if c[4]}


# ── Template generation ─────────────────────────────────────────────

def generate_employee_import_template() -> bytes:
    """Return bytes of an .xlsx template with headers and example rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    example_font = Font(italic=True, color="808080")
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Headers (row 1)
    for col_idx, name in enumerate(HEADER_NAMES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Date column indices (1-based)
    date_cols = {i+1 for i, c in enumerate(COLUMNS) if c[0] in ("Дата приёма", "Дата рождения")}

    # Example rows (rows 2-3)
    for row_offset, example in enumerate(EXAMPLE_ROWS):
        row_num = row_offset + 2
        for col_idx, val in enumerate(example, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = thin_border
            # Для дат — явно текстовый формат, чтобы Excel не конвертировал
            if col_idx in date_cols:
                cell.number_format = "@"  # text format

    # Auto-width
    for col_idx in range(1, len(HEADER_NAMES) + 1):
        max_len = len(HEADER_NAMES[col_idx - 1])
        for row in range(2, 4):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Date parsing ─────────────────────────────────────────────────────

def _parse_date(val):
    """Parse date from DD.MM.YYYY, YYYY-MM-DD string or datetime object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    val = str(val).strip()
    if not val:
        return None
    # DD.MM.YYYY
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", val)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    # YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", val):
        try:
            return date.fromisoformat(val)
        except ValueError:
            return None
    return None


# ── Salary parsing ───────────────────────────────────────────────────

def _parse_salary(val):
    """Parse salary: remove spaces, convert to Decimal."""
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    val = str(val).strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not val:
        return Decimal("0")
    try:
        return Decimal(val)
    except (InvalidOperation, ValueError):
        return None


# ── Import logic ─────────────────────────────────────────────────────

def import_employees_from_excel(file, company) -> dict:
    """
    Import employees from an uploaded .xlsx file.

    Returns: {'created': N, 'skipped': N, 'errors': [...]}
    """
    result = {"created": 0, "skipped": 0, "errors": []}

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        result["errors"].append(f"Не удалось открыть файл: {e}")
        return result

    ws = wb.active

    # Read headers from row 1 — build column index map
    header_map = {}  # col_idx (0-based) -> header name
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())):
        if cell:
            normalized = str(cell).strip()
            header_map[col_idx] = normalized

    if not header_map:
        result["errors"].append("Не найдены заголовки в первой строке")
        return result

    # Map header name -> column index
    name_to_col = {}
    for col_idx, name in header_map.items():
        name_to_col[name] = col_idx

    # Verify required headers exist
    missing_headers = REQUIRED_FIELDS - set(name_to_col.keys())
    if missing_headers:
        result["errors"].append(f"Отсутствуют обязательные колонки: {', '.join(sorted(missing_headers))}")
        return result

    # Build field mapping: field_name -> col_idx
    field_map = {}
    for header, field_name, _, _, _ in COLUMNS:
        if header in name_to_col:
            field_map[field_name] = name_to_col[header]

    # Example last names to skip (from template)
    example_last_names = {row[0] for row in EXAMPLE_ROWS}

    # Department cache
    dept_cache = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def get_val(field_name):
            idx = field_map.get(field_name)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            if isinstance(v, str):
                return v.strip() or None
            return v

        last_name = get_val("last_name")
        first_name = get_val("first_name")

        # Skip example rows
        if last_name in example_last_names and first_name in ("Иван", "Мария"):
            result["skipped"] += 1
            continue

        # Validate required fields
        position = get_val("position")
        hire_date_raw = get_val("hire_date")

        errors_in_row = []
        if not last_name:
            errors_in_row.append("Фамилия")
        if not first_name:
            errors_in_row.append("Имя")
        if not position:
            errors_in_row.append("Должность")
        if not hire_date_raw:
            errors_in_row.append("Дата приёма")

        if errors_in_row:
            result["errors"].append(f"Строка {row_num}: не заполнены — {', '.join(errors_in_row)}")
            result["skipped"] += 1
            continue

        # Parse hire_date
        hire_date = _parse_date(hire_date_raw)
        if not hire_date:
            result["errors"].append(f"Строка {row_num}: неверный формат даты приёма «{hire_date_raw}»")
            result["skipped"] += 1
            continue

        # Parse salary
        salary_raw = get_val("salary")
        salary = _parse_salary(salary_raw)
        if salary is None:
            result["errors"].append(f"Строка {row_num}: неверный формат оклада «{salary_raw}»")
            salary = Decimal("0")

        # Parse birth_date (optional)
        birth_date = _parse_date(get_val("birth_date"))

        # Department
        dept_name = get_val("department")
        department = None
        if dept_name:
            dept_name_str = str(dept_name)
            if dept_name_str not in dept_cache:
                dept_obj, _ = Department.objects.get_or_create(
                    company=company,
                    name=dept_name_str,
                )
                dept_cache[dept_name_str] = dept_obj
            department = dept_cache[dept_name_str]

        try:
            Employee.objects.create(
                company=company,
                last_name=str(last_name),
                first_name=str(first_name),
                middle_name=str(get_val("middle_name") or ""),
                position=str(position),
                hire_date=hire_date,
                salary=salary,
                phone=str(get_val("phone") or ""),
                email=str(get_val("email") or ""),
                department=department,
                inn=str(get_val("inn") or ""),
                snils=str(get_val("snils") or ""),
                birth_date=birth_date,
            )
            result["created"] += 1
        except Exception as e:
            result["errors"].append(f"Строка {row_num}: {e}")
            result["skipped"] += 1

    wb.close()
    return result
