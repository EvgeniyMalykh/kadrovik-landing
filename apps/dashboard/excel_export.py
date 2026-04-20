"""
Экспорт табеля учёта рабочего времени в Excel (.xlsx).
Два формата: свободная таблица и унифицированная форма Т-13.
"""
import calendar
import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.employees.models import Employee, TimeRecord


# ── Общие стили ──────────────────────────────────────────────────────────────

_thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Коды, которые считаются рабочими
_WORK_CODES = {'Я', 'К', 'Я½', 'РВ', 'Я/С'}


def _load_data(company, year, month):
    """Загрузить сотрудников, записи табеля, праздники и сокращённые дни."""
    days_in_month = calendar.monthrange(year, month)[1]
    employees = list(
        Employee.objects.filter(company=company)
        .select_related('department')
        .order_by('last_name')
    )
    start = datetime.date(year, month, 1)
    end = datetime.date(year, month, days_in_month)
    records = TimeRecord.objects.filter(
        employee__in=employees, date__gte=start, date__lte=end,
    )
    rec_map = {(r.employee_id, r.date.day): r for r in records}

    # Праздники
    from apps.documents.services import _get_ru_holidays
    holidays = _get_ru_holidays(year)

    # Сокращённые предпраздничные дни
    try:
        from apps.employees.utils import get_holidays_and_short_days
        _, short_days = get_holidays_and_short_days(year, month)
    except Exception:
        short_days = set()

    # Типы дней
    day_types = []
    for d in range(1, days_in_month + 1):
        dd = datetime.date(year, month, d)
        if dd in holidays:
            day_types.append('holiday')
        elif dd.weekday() >= 5:
            day_types.append('weekend')
        elif dd in short_days:
            day_types.append('short')
        else:
            day_types.append('work')

    return employees, rec_map, days_in_month, day_types, short_days, holidays


def _get_cell_data(rec, day_type):
    """Вернуть (code, hours) для ячейки на основе записи и типа дня."""
    if rec:
        code = rec.code
        hours = rec.hours
        if code == 'Я½' and not hours:
            hours = 4
        elif code in _WORK_CODES and not hours:
            hours = 8
        return code, hours

    if day_type == 'work':
        return 'Я', 8
    elif day_type == 'short':
        return 'Я', 7
    elif day_type == 'holiday':
        return 'П', 0
    else:
        return 'В', 0


_MONTH_NAMES = [
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Свободная таблица
# ═══════════════════════════════════════════════════════════════════════════════

def export_timesheet_free(company, year, month) -> bytes:
    """
    Свободная читаемая таблица табеля для бухгалтерии.
    Возвращает bytes Excel-файла.
    """
    employees, rec_map, days_in_month, day_types, short_days, holidays = \
        _load_data(company, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Табель {_MONTH_NAMES[month - 1]} {year}'

    # ── Стили ────────────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color='444444')
    header_font = Font(bold=True, size=9, color='000000')
    header_fill = PatternFill('solid', fgColor='D3D3D3')
    data_font = Font(size=9)
    bold_font = Font(bold=True, size=9)
    alt_fill = PatternFill('solid', fgColor='F5F5F5')
    weekend_fill = PatternFill('solid', fgColor='E8F4FD')

    total_cols = 3 + days_in_month + 2  # №, ФИО, Должность, дни..., Итого дней, Итого часов

    # ── Строка 1: заголовок ──────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, f'Табель учёта рабочего времени — {_MONTH_NAMES[month - 1]} {year}')
    c.font = title_font
    c.alignment = _center
    ws.row_dimensions[1].height = 24

    # ── Строка 2: компания ───────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws.cell(2, 1, company.name)
    c.font = subtitle_font
    c.alignment = _center
    ws.row_dimensions[2].height = 20

    # ── Строка 3: пустая ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Строка 4: шапка ──────────────────────────────────────────────────────
    hdr_row = 4
    headers = ['№', 'ФИО', 'Должность']
    col_widths = [5, 25, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(hdr_row, ci, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = _center
        c.border = _thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Дни месяца
    for d in range(1, days_in_month + 1):
        ci = 3 + d
        c = ws.cell(hdr_row, ci, d)
        c.font = header_font
        c.fill = header_fill
        c.alignment = _center
        c.border = _thin_border
        ws.column_dimensions[get_column_letter(ci)].width = 3.8

    # Итого
    ci_days = 3 + days_in_month + 1
    ci_hours = 3 + days_in_month + 2
    for ci, label in [(ci_days, 'Итого\nдней'), (ci_hours, 'Итого\nчасов')]:
        c = ws.cell(hdr_row, ci, label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = _center
        c.border = _thin_border
        ws.column_dimensions[get_column_letter(ci)].width = 8

    ws.row_dimensions[hdr_row].height = 28

    # ── Данные ───────────────────────────────────────────────────────────────
    for ri, emp in enumerate(employees):
        row_num = 5 + ri
        is_alt = ri % 2 == 1
        row_fill = alt_fill if is_alt else None

        # №
        c = ws.cell(row_num, 1, ri + 1)
        c.font = data_font
        c.alignment = _center
        c.border = _thin_border
        if row_fill:
            c.fill = row_fill

        # ФИО
        ln = emp.last_name or ''
        fn = emp.first_name or ''
        mn = emp.middle_name or ''
        full_name = f'{ln} {fn} {mn}'.strip()
        c = ws.cell(row_num, 2, full_name)
        c.font = data_font
        c.alignment = _left
        c.border = _thin_border
        if row_fill:
            c.fill = row_fill

        # Должность
        c = ws.cell(row_num, 3, emp.position or '')
        c.font = data_font
        c.alignment = _left
        c.border = _thin_border
        if row_fill:
            c.fill = row_fill

        total_days = 0
        total_hours = 0

        for d in range(1, days_in_month + 1):
            ci = 3 + d
            dtype = day_types[d - 1]
            rec = rec_map.get((emp.id, d))
            code, hours = _get_cell_data(rec, dtype)

            c = ws.cell(row_num, ci, code)
            c.font = Font(size=8)
            c.alignment = _center
            c.border = _thin_border

            # Фон: выходные/праздники — голубой
            if dtype in ('weekend', 'holiday') and code in ('В', 'П'):
                c.fill = weekend_fill
            elif is_alt:
                c.fill = alt_fill if row_fill else None

            if code in _WORK_CODES:
                total_days += 0.5 if code == 'Я½' else 1
                total_hours += hours

        # Итого
        c = ws.cell(row_num, ci_days, total_days)
        c.font = bold_font
        c.alignment = _center
        c.border = _thin_border
        if row_fill:
            c.fill = row_fill

        c = ws.cell(row_num, ci_hours, total_hours)
        c.font = bold_font
        c.alignment = _center
        c.border = _thin_border
        if row_fill:
            c.fill = row_fill

        ws.row_dimensions[row_num].height = 18

    # Заморозка панелей
    ws.freeze_panes = 'D5'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Форма Т-13
# ═══════════════════════════════════════════════════════════════════════════════

def export_timesheet_t13(company, year, month) -> bytes:
    """
    Унифицированная форма Т-13 по Постановлению Госкомстата РФ №1 от 05.01.2004.
    Возвращает bytes Excel-файла.
    """
    employees, rec_map, days_in_month, day_types, short_days, holidays = \
        _load_data(company, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Т-13'

    # ── Стили ────────────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=10)
    small_font = Font(size=8)
    small_bold = Font(bold=True, size=8)
    tiny_font = Font(size=7)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    weekend_fill = PatternFill('solid', fgColor='D9D9D9')
    border = _thin_border
    center = _center

    month_name = _MONTH_NAMES[month - 1]
    last_day = days_in_month

    # ── ШАПКА ────────────────────────────────────────────────────────────────
    # Строка 1: Организация
    ws.merge_cells('A1:F1')
    c = ws.cell(1, 1, company.name)
    c.font = subtitle_font
    c.alignment = Alignment(horizontal='left', vertical='center')

    # Унифицированная форма справа
    right_col = 4 + days_in_month + 6  # примерная правая граница
    if right_col > 10:
        ws.merge_cells(start_row=1, start_column=right_col - 5, end_row=1, end_column=right_col)
        c = ws.cell(1, right_col - 5, 'Унифицированная форма № Т-13')
        c.font = Font(size=8, italic=True)
        c.alignment = Alignment(horizontal='right', vertical='center')

    # Строка 2: Подразделение
    ws.merge_cells('A2:F2')
    c = ws.cell(2, 1, 'Структурное подразделение: ___________________')
    c.font = small_font
    c.alignment = Alignment(horizontal='left', vertical='center')

    # Строка 3: ОКПО
    ws.merge_cells('A3:F3')
    c = ws.cell(3, 1, f'ОКПО: {getattr(company, "okpo", "") or "________"}')
    c.font = small_font

    # Строка 4: ТАБЕЛЬ
    total_cols = 4 + days_in_month + 7  # №, ФИО, Таб.№, дни(код+часы), 1п дней, 1п часов, 2п дней, 2п часов, итого дней, итого часов, неявки
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    c = ws.cell(4, 1, 'ТАБЕЛЬ')
    c.font = Font(bold=True, size=16)
    c.alignment = _center
    ws.row_dimensions[4].height = 24

    # Строка 5: подзаголовок
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=total_cols)
    c = ws.cell(5, 1, 'учёта рабочего времени')
    c.font = Font(size=11)
    c.alignment = _center

    # Строка 6: номер, дата, период
    ws.merge_cells('A6:B6')
    ws.cell(6, 1, 'Номер документа').font = small_bold
    ws.cell(6, 1).alignment = _center
    ws.cell(6, 1).border = border
    ws.merge_cells('C6:D6')
    ws.cell(6, 3, 'Дата составления').font = small_bold
    ws.cell(6, 3).alignment = _center
    ws.cell(6, 3).border = border
    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=8)
    ws.cell(6, 5, 'Отчётный период').font = small_bold
    ws.cell(6, 5).alignment = _center
    ws.cell(6, 5).border = border

    ws.merge_cells('A7:B7')
    ws.cell(7, 1).border = border
    ws.merge_cells('C7:D7')
    today_str = datetime.date.today().strftime('%d.%m.%Y')
    ws.cell(7, 3, today_str).font = small_font
    ws.cell(7, 3).alignment = _center
    ws.cell(7, 3).border = border
    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=8)
    period_str = f'с 01.{month:02d}.{year} по {last_day:02d}.{month:02d}.{year}'
    ws.cell(7, 5, period_str).font = small_font
    ws.cell(7, 5).alignment = _center
    ws.cell(7, 5).border = border

    ws.row_dimensions[8].height = 6  # пустая разделительная строка

    # ── ЗАГОЛОВКИ ТАБЛИЦЫ ────────────────────────────────────────────────────
    # Две строки заголовков: строка 9 (верхний) и строка 10 (нижний)
    hdr1 = 9
    hdr2 = 10

    # Колонка 1: № п/п
    ws.merge_cells(start_row=hdr1, start_column=1, end_row=hdr2, end_column=1)
    c = ws.cell(hdr1, 1, '№\nп/п')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.column_dimensions['A'].width = 4

    # Колонка 2: ФИО, должность
    ws.merge_cells(start_row=hdr1, start_column=2, end_row=hdr2, end_column=2)
    c = ws.cell(hdr1, 2, 'Фамилия, инициалы,\nдолжность (специальность)')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.column_dimensions['B'].width = 28

    # Колонка 3: Табельный номер
    ws.merge_cells(start_row=hdr1, start_column=3, end_row=hdr2, end_column=3)
    c = ws.cell(hdr1, 3, 'Таб.\nномер')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.column_dimensions['C'].width = 6

    # Колонки дней: с колонки 4 по (4 + days_in_month - 1)
    # Верхняя строка (hdr1): "Отметки о явках и неявках на работу по числам месяца"
    day_start_col = 4
    day_end_col = day_start_col + days_in_month - 1
    ws.merge_cells(start_row=hdr1, start_column=day_start_col, end_row=hdr1, end_column=day_end_col)
    c = ws.cell(hdr1, day_start_col, 'Отметки о явках и неявках на работу по числам месяца')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border

    # Нижняя строка (hdr2): числа 1..31
    for d in range(1, days_in_month + 1):
        ci = day_start_col + d - 1
        dtype = day_types[d - 1]
        c = ws.cell(hdr2, ci, d)
        c.font = small_bold
        c.fill = weekend_fill if dtype in ('weekend', 'holiday') else header_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = 3.5

    # Колонки итогов
    sum_col = day_end_col + 1

    # Итого за 1-ю половину
    ws.merge_cells(start_row=hdr1, start_column=sum_col, end_row=hdr1, end_column=sum_col + 1)
    c = ws.cell(hdr1, sum_col, 'Отработано за\n1-ю половину')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.cell(hdr2, sum_col, 'дн.').font = small_bold
    ws.cell(hdr2, sum_col).fill = header_fill
    ws.cell(hdr2, sum_col).alignment = center
    ws.cell(hdr2, sum_col).border = border
    ws.column_dimensions[get_column_letter(sum_col)].width = 5
    ws.cell(hdr2, sum_col + 1, 'час.').font = small_bold
    ws.cell(hdr2, sum_col + 1).fill = header_fill
    ws.cell(hdr2, sum_col + 1).alignment = center
    ws.cell(hdr2, sum_col + 1).border = border
    ws.column_dimensions[get_column_letter(sum_col + 1)].width = 5

    # Итого за 2-ю половину
    ws.merge_cells(start_row=hdr1, start_column=sum_col + 2, end_row=hdr1, end_column=sum_col + 3)
    c = ws.cell(hdr1, sum_col + 2, 'Отработано за\n2-ю половину')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.cell(hdr2, sum_col + 2, 'дн.').font = small_bold
    ws.cell(hdr2, sum_col + 2).fill = header_fill
    ws.cell(hdr2, sum_col + 2).alignment = center
    ws.cell(hdr2, sum_col + 2).border = border
    ws.column_dimensions[get_column_letter(sum_col + 2)].width = 5
    ws.cell(hdr2, sum_col + 3, 'час.').font = small_bold
    ws.cell(hdr2, sum_col + 3).fill = header_fill
    ws.cell(hdr2, sum_col + 3).alignment = center
    ws.cell(hdr2, sum_col + 3).border = border
    ws.column_dimensions[get_column_letter(sum_col + 3)].width = 5

    # Итого за месяц
    ws.merge_cells(start_row=hdr1, start_column=sum_col + 4, end_row=hdr1, end_column=sum_col + 5)
    c = ws.cell(hdr1, sum_col + 4, 'Итого за\nмесяц')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.cell(hdr2, sum_col + 4, 'дн.').font = small_bold
    ws.cell(hdr2, sum_col + 4).fill = header_fill
    ws.cell(hdr2, sum_col + 4).alignment = center
    ws.cell(hdr2, sum_col + 4).border = border
    ws.column_dimensions[get_column_letter(sum_col + 4)].width = 5
    ws.cell(hdr2, sum_col + 5, 'час.').font = small_bold
    ws.cell(hdr2, sum_col + 5).fill = header_fill
    ws.cell(hdr2, sum_col + 5).alignment = center
    ws.cell(hdr2, sum_col + 5).border = border
    ws.column_dimensions[get_column_letter(sum_col + 5)].width = 6

    # Неявки (ОТ, Б)
    ws.merge_cells(start_row=hdr1, start_column=sum_col + 6, end_row=hdr1, end_column=sum_col + 9)
    c = ws.cell(hdr1, sum_col + 6, 'Неявки по причинам')
    c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border

    absence_hdrs = ['ОТ', 'дн.', 'Б', 'дн.']
    for i, label in enumerate(absence_hdrs):
        ci = sum_col + 6 + i
        c = ws.cell(hdr2, ci, label)
        c.font = small_bold; c.fill = header_fill; c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = 4.5

    ws.row_dimensions[hdr1].height = 32
    ws.row_dimensions[hdr2].height = 18

    # ── ДАННЫЕ СОТРУДНИКОВ ───────────────────────────────────────────────────
    # Каждый сотрудник = 2 строки: верхняя — код, нижняя — часы
    data_start = 11
    mid_day = 15  # первая половина: 1..15, вторая: 16..last

    for emp_idx, emp in enumerate(employees):
        row_code = data_start + emp_idx * 2
        row_hrs = row_code + 1

        # №
        ws.merge_cells(start_row=row_code, start_column=1, end_row=row_hrs, end_column=1)
        c = ws.cell(row_code, 1, emp_idx + 1)
        c.font = small_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, 1).border = border

        # ФИО + должность
        ln = emp.last_name or ''
        fn = (emp.first_name[:1] + '.') if emp.first_name else ''
        mn = (emp.middle_name[:1] + '.') if emp.middle_name else ''
        short_name = f'{ln} {fn}{mn}'.strip()
        pos = emp.position or ''

        ws.merge_cells(start_row=row_code, start_column=2, end_row=row_hrs, end_column=2)
        c = ws.cell(row_code, 2, f'{short_name}\n{pos}')
        c.font = small_font; c.alignment = _left; c.border = border
        ws.cell(row_hrs, 2).border = border

        # Табельный номер
        tab_num = getattr(emp, 'personnel_number', '') or str(emp.id)
        ws.merge_cells(start_row=row_code, start_column=3, end_row=row_hrs, end_column=3)
        c = ws.cell(row_code, 3, tab_num)
        c.font = small_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, 3).border = border

        # Подсчёт по половинам
        half1_days = 0
        half1_hours = 0
        half2_days = 0
        half2_hours = 0
        total_days = 0
        total_hours = 0
        absence_ot = 0  # дней отпуска
        absence_b = 0   # дней больничного

        for d in range(1, days_in_month + 1):
            ci = day_start_col + d - 1
            dtype = day_types[d - 1]
            rec = rec_map.get((emp.id, d))
            code, hours = _get_cell_data(rec, dtype)

            # Код (верхняя строка)
            c = ws.cell(row_code, ci, code)
            c.font = tiny_font; c.alignment = center; c.border = border
            if dtype in ('weekend', 'holiday') and code in ('В', 'П'):
                c.fill = weekend_fill

            # Часы (нижняя строка)
            c = ws.cell(row_hrs, ci, hours if hours else '')
            c.font = tiny_font; c.alignment = center; c.border = border
            if dtype in ('weekend', 'holiday') and code in ('В', 'П'):
                c.fill = weekend_fill

            # Подсчёт
            if code in _WORK_CODES:
                dd = 0.5 if code == 'Я½' else 1
                if d <= mid_day:
                    half1_days += dd
                    half1_hours += hours
                else:
                    half2_days += dd
                    half2_hours += hours
                total_days += dd
                total_hours += hours

            # Неявки
            if code == 'ОТ' or code == 'ОД' or code == 'УЧ' or code == 'ОЖ':
                absence_ot += 1
            elif code == 'Б':
                absence_b += 1

        # Итого за 1-ю половину
        ws.merge_cells(start_row=row_code, start_column=sum_col, end_row=row_hrs, end_column=sum_col)
        c = ws.cell(row_code, sum_col, half1_days)
        c.font = small_bold; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col).border = border

        ws.merge_cells(start_row=row_code, start_column=sum_col + 1, end_row=row_hrs, end_column=sum_col + 1)
        c = ws.cell(row_code, sum_col + 1, half1_hours)
        c.font = small_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 1).border = border

        # Итого за 2-ю половину
        ws.merge_cells(start_row=row_code, start_column=sum_col + 2, end_row=row_hrs, end_column=sum_col + 2)
        c = ws.cell(row_code, sum_col + 2, half2_days)
        c.font = small_bold; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 2).border = border

        ws.merge_cells(start_row=row_code, start_column=sum_col + 3, end_row=row_hrs, end_column=sum_col + 3)
        c = ws.cell(row_code, sum_col + 3, half2_hours)
        c.font = small_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 3).border = border

        # Итого за месяц
        ws.merge_cells(start_row=row_code, start_column=sum_col + 4, end_row=row_hrs, end_column=sum_col + 4)
        c = ws.cell(row_code, sum_col + 4, total_days)
        c.font = small_bold; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 4).border = border

        ws.merge_cells(start_row=row_code, start_column=sum_col + 5, end_row=row_hrs, end_column=sum_col + 5)
        c = ws.cell(row_code, sum_col + 5, total_hours)
        c.font = small_bold; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 5).border = border

        # Неявки: ОТ
        ws.merge_cells(start_row=row_code, start_column=sum_col + 6, end_row=row_hrs, end_column=sum_col + 6)
        c = ws.cell(row_code, sum_col + 6, 'ОТ' if absence_ot else '')
        c.font = tiny_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 6).border = border

        ws.merge_cells(start_row=row_code, start_column=sum_col + 7, end_row=row_hrs, end_column=sum_col + 7)
        c = ws.cell(row_code, sum_col + 7, absence_ot if absence_ot else '')
        c.font = small_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 7).border = border

        # Неявки: Б
        ws.merge_cells(start_row=row_code, start_column=sum_col + 8, end_row=row_hrs, end_column=sum_col + 8)
        c = ws.cell(row_code, sum_col + 8, 'Б' if absence_b else '')
        c.font = tiny_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 8).border = border

        ws.merge_cells(start_row=row_code, start_column=sum_col + 9, end_row=row_hrs, end_column=sum_col + 9)
        c = ws.cell(row_code, sum_col + 9, absence_b if absence_b else '')
        c.font = small_font; c.alignment = center; c.border = border
        ws.cell(row_hrs, sum_col + 9).border = border

        ws.row_dimensions[row_code].height = 16
        ws.row_dimensions[row_hrs].height = 14

    # ── ПОДПИСИ ──────────────────────────────────────────────────────────────
    sign_row = data_start + len(employees) * 2 + 2
    sign_font = Font(size=9)

    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=10)
    ws.cell(sign_row, 1, 'Руководитель  ________________ / ______________________ /').font = sign_font

    ws.merge_cells(start_row=sign_row + 2, start_column=1, end_row=sign_row + 2, end_column=10)
    ws.cell(sign_row + 2, 1, 'Работник кадровой службы  ________________ / ______________________ /').font = sign_font

    ws.merge_cells(start_row=sign_row + 4, start_column=1, end_row=sign_row + 4, end_column=10)
    ws.cell(sign_row + 4, 1, 'Бухгалтер  ________________ / ______________________ /').font = sign_font

    # Заморозка
    ws.freeze_panes = ws.cell(data_start, day_start_col).coordinate

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
