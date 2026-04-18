import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from apps.accounts.models import User
from apps.employees.models import Employee, Department
from apps.companies.models import Company, CompanyMember
from apps.documents.services import generate_t1_pdf, generate_t2_pdf, generate_t8_pdf, generate_t6_pdf
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from functools import wraps
from django.contrib import messages
from django.conf import settings

# ===== ROLE-BASED ACCESS CONTROL =====

ROLE_RANK = {
    'owner': 4,
    'admin': 3,
    'hr': 2,
    'accountant': 1,
}

def get_member_role(user):
    """Получить роль текущего пользователя в его компании."""
    member = CompanyMember.objects.filter(user=user).first()
    return member.role if member else None


def require_role(min_role):
    """
    Декоратор: требует роль не ниже min_role.
    Порядок: owner > admin > hr > accountant

    Использование:
        @require_role('admin')  # нужна роль admin или выше
        def some_view(request): ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect(settings.LOGIN_URL + '?next=' + request.path)
            role = get_member_role(request.user)
            if not role or ROLE_RANK.get(role, 0) < ROLE_RANK.get(min_role, 99):
                messages.error(request, 'У вас недостаточно прав для этого действия.')
                return redirect('dashboard:home')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator



@login_required
def dashboard_home(request):
    return redirect("dashboard:employees")


def subscription_required(view_func):
    """Декоратор: проверяет активную подписку. При истечении — редирект на тарифы."""
    from functools import wraps
    from django.contrib import messages
    from django.shortcuts import redirect
    from apps.billing.models import Subscription as _Sub
    from apps.companies.models import CompanyMember as _CM
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        member = _CM.objects.filter(user=request.user).first()
        if not member:
            return view_func(request, *args, **kwargs)
        sub = getattr(member.company, "subscription", None)
        if sub and not sub.is_active:
            messages.error(request, "Подписка истекла. Выберите тариф для продолжения работы.")
            return redirect("dashboard:subscription")
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def employees_list(request):
    member = CompanyMember.objects.filter(user=request.user).first()
    if member:
        company = member.company
    else:
        return redirect('dashboard:login')

    from datetime import date as _date
    from apps.billing.models import Subscription
    from django.utils import timezone
    import datetime as _dt
    from django.db.models import Q

    sub = getattr(company, "subscription", None) if company else None
    # Для старых аккаунтов без подписки — создаём trial
    if company and not sub:
        sub = Subscription.objects.create(
            company=company,
            plan=Subscription.Plan.TRIAL,
            status=Subscription.Status.ACTIVE,
            expires_at=timezone.now() + _dt.timedelta(days=7),
            max_employees=50,
        )
    # Оставшиеся дни trial
    trial_days_left = None
    if sub and sub.plan == Subscription.Plan.TRIAL and sub.expires_at:
        delta = sub.expires_at - timezone.now()
        trial_days_left = max(0, delta.days)
    from apps.billing.services import get_subscription_context
    sub_ctx = get_subscription_context(company)
    # Если подписка истекла — рендерим страницу с флагом (не редирект, т.к. главная)
    subscription_expired = bool(sub and not sub.is_active)

    # Параметры фильтрации
    q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'active')
    department_filter = request.GET.get('department', '')
    contract_filter = request.GET.get('contract', '')

    employees = Employee.objects.filter(company=company).select_related('department')

    # Фильтр по статусу
    if status_filter == 'active':
        employees = employees.filter(status='active')
    elif status_filter == 'fired':
        employees = employees.filter(status='fired')
    # 'all' — без фильтра по статусу

    # Поиск по ФИО, должности, отделу, таб. номеру, телефону
    if q:
        employees = employees.filter(
            Q(last_name__icontains=q) |
            Q(first_name__icontains=q) |
            Q(middle_name__icontains=q) |
            Q(position__icontains=q) |
            Q(department__name__icontains=q) |
            Q(personnel_number__icontains=q) |
            Q(phone__icontains=q)
        )

    # Фильтр по отделу
    if department_filter:
        employees = employees.filter(department_id=department_filter)

    # Фильтр по типу договора
    if contract_filter:
        employees = employees.filter(contract_type=contract_filter)

    employees = employees.order_by('last_name', 'first_name')

    # Счётчики для вкладок
    active_count = Employee.objects.filter(company=company, status='active').count()
    fired_count = Employee.objects.filter(company=company, status='fired').count()

    # Отделы для фильтра
    departments = Department.objects.filter(company=company).order_by('name')

    return render(request, "dashboard/employees.html", {
        "employees": employees,
        "company": company,
        "today": _date.today().isoformat(),
        "sub": sub,
        "trial_days_left": trial_days_left,
        "subscription_expired": subscription_expired,
        "q": q,
        "status_filter": status_filter,
        "department_filter": department_filter,
        "contract_filter": contract_filter,
        "active_count": active_count,
        "fired_count": fired_count,
        "departments": departments,
        "total_count": employees.count(),
        **sub_ctx,
    })





def _parse_date_flexible(val):
    """Parse date from DD.MM.YYYY or YYYY-MM-DD format. Returns date or None."""
    if not val or not val.strip():
        return None
    val = val.strip()
    # Try DD.MM.YYYY
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", val)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    # Try YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", val):
        try:
            return date.fromisoformat(val)
        except ValueError:
            return None
    return None


def _save_employee_from_post(post, employee):
    """Обновляет поля сотрудника из POST-данных."""
    is_new = not employee.pk

    employee.last_name   = post.get("last_name", "")
    employee.first_name  = post.get("first_name", "")
    employee.middle_name = post.get("middle_name", "")
    employee.position    = post.get("position", "")
    salary_raw = post.get('salary')
    if salary_raw not in (None, ''):
        try:
            employee.salary = Decimal(salary_raw)
        except (InvalidOperation, ValueError):
            employee.salary = None
    else:
        employee.salary = None

    # Структурное подразделение
    dept_id = post.get("department_id")
    dept_name_new = post.get("department_new", "").strip()
    if dept_name_new and hasattr(employee, 'company') and employee.company:
        from apps.employees.models import Department as _Dept
        dept, _ = _Dept.objects.get_or_create(company=employee.company, name=dept_name_new)
        employee.department = dept
    elif dept_id:
        from apps.employees.models import Department as _Dept
        try:
            employee.department = _Dept.objects.get(id=int(dept_id))
        except (_Dept.DoesNotExist, ValueError):
            employee.department = None
    else:
        employee.department = None

    # Даты: при редактировании пустое поле не затирает старое значение
    parsed_hire = _parse_date_flexible(post.get("hire_date"))
    if parsed_hire:
        employee.hire_date = parsed_hire
    elif is_new:
        employee.hire_date = date.today()

    parsed_birth = _parse_date_flexible(post.get("birth_date"))
    if parsed_birth:
        employee.birth_date = parsed_birth
    elif is_new:
        employee.birth_date = None

    probation_months = post.get("probation_months")
    if probation_months:
        try:
            months = int(probation_months)
            employee.probation_end_date = employee.hire_date + timedelta(days=30 * months)
        except ValueError:
            if is_new:
                employee.probation_end_date = None
    else:
        probation_end_str = post.get("probation_end_date")
        parsed_probation = _parse_date_flexible(probation_end_str)
        if parsed_probation:
            employee.probation_end_date = parsed_probation
        elif is_new:
            employee.probation_end_date = None

    employee.phone  = post.get("phone", "")
    employee.email  = post.get("email", "")
    employee.inn    = post.get("inn", "")
    employee.snils  = post.get("snils", "")
    employee.passport_series        = post.get("passport_series", "")
    employee.passport_number        = post.get("passport_number", "")
    employee.passport_issued_by     = post.get("passport_issued_by", "")
    employee.passport_registration  = post.get("passport_registration", "")
    employee.personnel_number       = post.get("personnel_number", "")
    employee.status                 = post.get("status", "active")
    employee.contract_type          = post.get("contract_type", "permanent")

    parsed_passport_date = _parse_date_flexible(post.get("passport_issued_date"))
    if parsed_passport_date:
        employee.passport_issued_date = parsed_passport_date
    elif is_new:
        employee.passport_issued_date = None

    parsed_contract_end = _parse_date_flexible(post.get("contract_end_date"))
    if parsed_contract_end:
        employee.contract_end_date = parsed_contract_end
    elif is_new:
        employee.contract_end_date = None

    parsed_fire = _parse_date_flexible(post.get("fire_date"))
    if parsed_fire:
        employee.fire_date = parsed_fire
    elif is_new:
        employee.fire_date = None

    # Новые поля
    employee.birth_place = post.get("birth_place", "")
    employee.education   = post.get("education", "")
    employee.marital_status = post.get("marital_status", "") or ""
    employee.citizenship    = post.get("citizenship", "") or "Российская Федерация"

    return employee


@login_required
@subscription_required
@require_role("hr")
def employee_add(request):
    from apps.billing.services import get_subscription_context
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return HttpResponse("Нет компании", status=400)
    sub_ctx = get_subscription_context(member.company)

    if request.method == "POST":
        # Проверяем лимит сотрудников по тарифу
        if not sub_ctx["can_add_employee"]:
            return HttpResponse(
                f'<div class="alert alert-error" style="padding:12px 16px;border-radius:8px;background:#fee2e2;color:#991b1b;margin:8px 0;">'
                f'Достигнут лимит тарифа — <strong>{sub_ctx["max_employees"]} сотрудников</strong>. '
                f'<a href="/dashboard/subscription/" style="color:#991b1b;font-weight:600;">Обновите тариф</a> для добавления новых.</div>',
                status=200
            )
        emp = Employee(company=member.company)
        _save_employee_from_post(request.POST, emp)
        emp.save()
        employees = Employee.objects.filter(company=member.company).select_related("department")
        return render(request, "dashboard/partials/employees_table.html", {"employees": employees})

    departments = list(member.company.departments.values('id', 'name')) if hasattr(member.company, 'departments') else []
    from apps.employees.models import Department as _Dept
    departments = list(_Dept.objects.filter(company=member.company).values('id', 'name'))
    # Автотабельный номер
    import re as _re_pn
    existing_nums = Employee.objects.filter(company=member.company).values_list('personnel_number', flat=True)
    max_pn = 0
    for pn in existing_nums:
        if pn:
            m = _re_pn.search(r'(\d+)', str(pn))
            if m:
                max_pn = max(max_pn, int(m.group(1)))
    next_personnel_number = str(max_pn + 1).zfill(3)
    return render(request, "dashboard/partials/employee_form.html", {
        "departments": departments,
        "can_add_employee": sub_ctx["can_add_employee"],
        "max_employees": sub_ctx["max_employees"],
        "employee_count": sub_ctx["employee_count"],
        "next_personnel_number": next_personnel_number,
    })


@login_required
@subscription_required
def employee_edit(request, employee_id):
    if not request.headers.get("HX-Request"):
        return redirect("dashboard:employees")
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    if request.method == "POST":
        role = get_member_role(request.user)
        if not role or ROLE_RANK.get(role, 0) < ROLE_RANK.get('hr', 0):
            messages.error(request, 'Недостаточно прав для редактирования.')
            return redirect('dashboard:employees')
        _save_employee_from_post(request.POST, employee)
        employee.save()
        employees = Employee.objects.filter(company=member.company).select_related("department")
        return render(request, "dashboard/partials/employees_table.html", {"employees": employees})
    from apps.employees.models import Department as _Dept
    departments = list(_Dept.objects.filter(company=member.company).values('id', 'name'))
    return render(request, "dashboard/partials/employee_edit_form.html", {
        "emp": employee,
        "birth_date_str":           employee.birth_date.strftime("%d.%m.%Y") if employee.birth_date else "",
        "hire_date_str":            employee.hire_date.strftime("%d.%m.%Y") if employee.hire_date else "",
        "probation_end_str":        employee.probation_end_date.strftime("%d.%m.%Y") if employee.probation_end_date else "",
        "contract_end_str":         employee.contract_end_date.strftime("%d.%m.%Y") if employee.contract_end_date else "",
        "fire_date_str":            employee.fire_date.strftime("%d.%m.%Y") if employee.fire_date else "",
        "passport_issued_date_str": employee.passport_issued_date.strftime("%d.%m.%Y") if employee.passport_issued_date else "",
        "departments": departments,
    })


@login_required
@require_POST
@require_role("admin")
def employee_delete(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    employee.delete()
    from datetime import date as _date
    employees = Employee.objects.filter(company=member.company).select_related("department")
    return render(request, "dashboard/partials/employees_table.html", {
        "employees": employees,
        "today": _date.today().strftime("%Y-%m-%d"),
    })


@login_required
@subscription_required
def download_t1(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    pdf = generate_t1_pdf(employee, request.GET.get("order", "П-001"))
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"T1_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_t2(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    pdf = generate_t2_pdf(employee)
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"T2_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_t8(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    pdf = generate_t8_pdf(employee, request.GET.get("order", "У-001"))
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"T8_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_t6(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from datetime import date as dt
    v_start = dt.fromisoformat(request.GET.get("start", dt.today().isoformat()))
    v_end   = dt.fromisoformat(request.GET.get("end",   dt.today().isoformat()))
    pdf = generate_t6_pdf(employee, v_start, v_end, request.GET.get("order", "О-001"), vacation_type=request.GET.get("vtype"))
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"T6_{employee.last_name}.pdf\""
    return r


@login_required
def subscription(request):
    member = CompanyMember.objects.filter(user=request.user).first()
    sub = None
    payments = []
    employee_count = 0
    if member:
        sub = getattr(member.company, "subscription", None)
        payments = member.company.payments.order_by('-created_at')[:10]
        employee_count = Employee.objects.filter(company=member.company).count()
    return render(request, "dashboard/subscription.html", {
        "sub": sub,
        "payments": payments,
        "employee_count": employee_count,
    })


def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard:employees")
    if request.method == "POST":
        email    = request.POST.get("email")
        password = request.POST.get("password")
        user = authenticate(request, username=email, password=password)
        if user:
            login(request, user)
            if request.POST.get("remember_me"):
                request.session.set_expiry(60 * 60 * 24 * 30)  # 30 дней
            else:
                request.session.set_expiry(0)  # до закрытия браузера
            return redirect("dashboard:employees")
        return render(request, "dashboard/login.html", {"error": "Неверный email или пароль"})
    return render(request, "dashboard/login.html")


def register_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard:employees")
    if request.method == "POST":
        email        = request.POST.get("email", "").strip().lower()
        password     = request.POST.get("password")
        password2    = request.POST.get("password2")
        company_name = request.POST.get("company_name", "").strip()

        if not email or not password or not company_name:
            return render(request, "dashboard/register.html", {"error": "Заполните все поля"})
        if password != password2:
            return render(request, "dashboard/register.html", {"error": "Пароли не совпадают"})
        if User.objects.filter(email=email).exists():
            return render(request, "dashboard/register.html", {"error": "Email уже зарегистрирован"})

        import uuid, hashlib
        from django.contrib.auth.hashers import make_password
        from django.utils import timezone
        import datetime

        # Сохраняем данные регистрации в Redis (не в БД!) до подтверждения email
        token = str(uuid.uuid4())
        expires_at = (timezone.now() + datetime.timedelta(hours=24)).isoformat()
        password_hash = make_password(password)

        import redis as _redis
        from django.conf import settings
        r = _redis.from_url(getattr(settings, "REDIS_RELAY_URL", "redis://redis:6379/2"))
        import json as _json
        pending_key = f"pending_registration:{token}"
        r.setex(pending_key, 86400, _json.dumps({
            "email": email,
            "password_hash": password_hash,
            "company_name": company_name,
            "expires_at": expires_at,
        }, ensure_ascii=False))

        verify_url = request.build_absolute_uri(f"/dashboard/verify-email/{token}/")

        # Отправляем только письмо — в БД ничего не создаётся
        from apps.accounts.tasks import send_verification_email_pending
        send_verification_email_pending.delay(email, verify_url)

        return render(request, "dashboard/email_sent.html", {"email": email})
    return render(request, "dashboard/register.html")


def verify_email_view(request, token):
    import redis as _redis, json as _json
    from django.utils import timezone
    from django.conf import settings
    import datetime

    r = _redis.from_url(getattr(settings, "REDIS_RELAY_URL", "redis://redis:6379/2"))
    pending_key = f"pending_registration:{token}"
    data_raw = r.get(pending_key)

    if not data_raw:
        return render(request, "dashboard/register.html", {"error": "Ссылка недействительна или устарела"})

    data = _json.loads(data_raw)
    email        = data["email"]
    password_hash = data["password_hash"]
    company_name  = data["company_name"]

    # Проверяем — вдруг успели зарегистрироваться с тем же email
    if User.objects.filter(email=email).exists():
        r.delete(pending_key)
        return render(request, "dashboard/register.html", {"error": "Email уже зарегистрирован"})

    # Создаём пользователя, компанию, подписку — только сейчас
    from apps.billing.models import Subscription

    user = User(username=email, email=email, email_verified=True, is_active=True)
    user.password = password_hash
    user.save()

    company = Company.objects.create(name=company_name, owner=user)
    CompanyMember.objects.create(user=user, company=company, role="owner")
    Subscription.objects.create(
        company=company,
        plan=Subscription.Plan.TRIAL,
        status=Subscription.Status.ACTIVE,
        expires_at=timezone.now() + datetime.timedelta(days=7),
        max_employees=10,
    )

    # Удаляем pending запись
    r.delete(pending_key)

    # Уведомления — Telegram + Google Sheets (только после реальной регистрации)
    from apps.accounts.tasks import notify_new_registration
    notify_new_registration.delay(
        email=email,
        company_name=company_name,
        registered_at=timezone.now().strftime("%d.%m.%Y %H:%M"),
    )

    login(request, user, backend="django.contrib.auth.backends.ModelBackend")
    return render(request, "dashboard/email_verified.html")


def resend_verification_view(request):
    import redis as _redis, json as _json, uuid
    from django.utils import timezone
    from django.conf import settings
    import datetime

    email = request.GET.get("email", "").strip().lower()
    if not email:
        return redirect("dashboard:register")

    r = _redis.from_url(getattr(settings, "REDIS_RELAY_URL", "redis://redis:6379/2"))

    # Ищем pending запись по email
    existing_data = None
    existing_key = None
    for key in r.scan_iter("pending_registration:*"):
        raw = r.get(key)
        if raw:
            d = _json.loads(raw)
            if d.get("email") == email:
                existing_data = d
                existing_key = key
                break

    if not existing_data:
        return redirect("dashboard:register")

    # Создаём новый токен
    new_token = str(uuid.uuid4())
    new_key = f"pending_registration:{new_token}"
    r.setex(new_key, 86400, _json.dumps(existing_data, ensure_ascii=False))
    if existing_key:
        r.delete(existing_key)

    verify_url = request.build_absolute_uri(f"/dashboard/verify-email/{new_token}/")
    from apps.accounts.tasks import send_verification_email_pending
    send_verification_email_pending.delay(email, verify_url)

    return render(request, "dashboard/email_sent.html", {"email": email, "resent": True})


def logout_view(request):
    logout(request)
    return redirect("dashboard:login")


@login_required
@subscription_required
def download_t5(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_t5_pdf
    new_position = request.GET.get("position", employee.position or "Новая должность")
    new_salary   = request.GET.get("salary")
    pdf = generate_t5_pdf(employee, new_position, new_salary, request.GET.get("order", "ПР-001"))
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"T5_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_salary_change(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_salary_change_pdf
    new_salary = request.GET.get("salary", str(employee.salary or "0"))
    # Прежний оклад: сначала из документа (надёжный источник), потом из GET-параметра
    old_salary = ""
    from apps.documents.models import Document
    last_doc = Document.objects.filter(
        employee=employee, doc_type='salary_change'
    ).order_by('-date', '-id').first()
    if last_doc and last_doc.extra_data and last_doc.extra_data.get('old_salary'):
        old_salary = last_doc.extra_data['old_salary']
    if not old_salary:
        old_salary = request.GET.get("old_salary", "")
    pdf = generate_salary_change_pdf(employee, new_salary, request.GET.get("order", "З-001"), previous_salary=old_salary)
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"SalaryChange_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_transfer_order(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_transfer_order_pdf
    new_position = request.GET.get("new_position", "")
    new_salary = request.GET.get("new_salary", "")
    transfer_date = request.GET.get("transfer_date", "")
    reason = request.GET.get("reason", "")
    pdf = generate_transfer_order_pdf(
        employee,
        new_position=new_position or (employee.position or "Новая должность"),
        new_salary=new_salary or None,
        order_number=request.GET.get("order", "ПР-001"),
        transfer_date=transfer_date or None,
        reason=reason or None,
    )
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"Transfer_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_dismissal_order(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_dismissal_order_pdf
    pdf = generate_dismissal_order_pdf(
        employee,
        order_number=request.GET.get("order", "У-001"),
        dismissal_date=request.GET.get("dismissal_date", "") or None,
        dismissal_reason=request.GET.get("dismissal_reason", "") or None,
        dismissal_basis_doc=request.GET.get("dismissal_basis_doc", "") or None,
    )
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"Dismissal_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_bonus_order(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_bonus_order_pdf
    pdf = generate_bonus_order_pdf(
        employee,
        bonus_amount=request.GET.get("bonus_amount", "0"),
        order_number=request.GET.get("order", "П-001"),
        reason=request.GET.get("reason", "") or None,
        payment_date=request.GET.get("payment_date", "") or None,
    )
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"Bonus_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_disciplinary_order(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_disciplinary_order_pdf
    pdf = generate_disciplinary_order_pdf(
        employee,
        penalty_type=request.GET.get("penalty_type", "Выговор"),
        order_number=request.GET.get("order", "ДВ-001"),
        violation_date=request.GET.get("violation_date", "") or None,
        violation_description=request.GET.get("violation_description", "") or None,
        reason=request.GET.get("reason", "") or None,
    )
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"Disciplinary_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_work_certificate(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_work_certificate_pdf
    pdf = generate_work_certificate_pdf(employee)
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"Certificate_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_labor_contract(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_labor_contract_pdf
    pdf = generate_labor_contract_pdf(employee)
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"LaborContract_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_gph_contract(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_gph_contract_pdf
    pdf = generate_gph_contract_pdf(employee)
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"GPH_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_gph_act(request, employee_id):
    member = CompanyMember.objects.filter(user=request.user).first()
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    from apps.documents.services import generate_gph_act_pdf
    pdf = generate_gph_act_pdf(employee)
    r = HttpResponse(pdf, content_type="application/pdf")
    r["Content-Disposition"] = f"attachment; filename=\"GPH_Act_{employee.last_name}.pdf\""
    return r


@login_required
@subscription_required
def download_t13(request):
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return HttpResponse("Нет компании", status=400)
    from apps.documents.services import generate_t13_pdf
    employees = list(Employee.objects.filter(company=member.company))
    if not employees:
        return HttpResponse("Нет сотрудников", status=400)
    from datetime import date
    today = date.today()
    try:
        y = int(request.GET.get("year", today.year))
        m = int(request.GET.get("month", today.month))
        if not (1 <= m <= 12): m = today.month
        if not (2000 <= y <= 2100): y = today.year
    except (ValueError, TypeError):
        y, m = today.year, today.month
    pdf = generate_t13_pdf(employees, y, m)
    r = HttpResponse(pdf, content_type="application/pdf")
    fname = "%04d_%02d" % (y, m)
    r["Content-Disposition"] = "attachment; filename=T13_" + fname + ".pdf"
    return r


@login_required
def company_profile(request):
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect("dashboard:employees")
    company = member.company
    saved = False
    if request.method == "POST":
        role = get_member_role(request.user)
        if not role or ROLE_RANK.get(role, 0) < ROLE_RANK.get('admin', 0):
            messages.error(request, 'Недостаточно прав для изменения настроек компании.')
            return redirect('dashboard:company')
        company.name             = request.POST.get("name", company.name)
        company.inn              = request.POST.get("inn", company.inn)
        company.ogrn             = request.POST.get("ogrn", company.ogrn)
        company.kpp              = request.POST.get("kpp", company.kpp)
        company.okpo             = request.POST.get("okpo", company.okpo)
        company.sfr_reg_number   = request.POST.get("sfr_reg_number", company.sfr_reg_number)
        company.okved            = request.POST.get("okved", company.okved)
        company.legal_address    = request.POST.get("legal_address", company.legal_address)
        company.actual_address   = request.POST.get("actual_address", company.actual_address)
        company.director_name    = request.POST.get("director_name", company.director_name)
        company.director_position = request.POST.get("director_position", company.director_position)
        company.phone            = request.POST.get("phone", company.phone)
        company.email            = request.POST.get("email", company.email)
        company.save()
        saved = True
    return render(request, "dashboard/company.html", {"company": company, "saved": saved})


@login_required
@subscription_required
def timesheet_edit(request):
    """Редактирование табеля Т-13 по месяцу."""
    import calendar
    from datetime import date as dt_date
    from apps.employees.models import TimeRecord
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect("dashboard:employees")
    try:
        y = int(request.GET.get("year", dt_date.today().year))
        m = int(request.GET.get("month", dt_date.today().month))
    except (ValueError, TypeError):
        y, m = dt_date.today().year, dt_date.today().month

    employees = Employee.objects.filter(company=member.company).select_related("department")
    days_in_month = calendar.monthrange(y, m)[1]
    days = list(range(1, days_in_month + 1))

    # Получаем все отметки за месяц
    from django.db.models import Q
    import datetime
    start = datetime.date(y, m, 1)
    end = datetime.date(y, m, days_in_month)
    records = TimeRecord.objects.filter(
        employee__in=employees,
        date__gte=start, date__lte=end
    )
    # Словарь: {(employee_id, day): record}
    rec_map = {(r.employee_id, r.date.day): r for r in records}

    # Праздники, сокращённые и выходные
    holidays = _get_ru_holidays_dashboard(y)
    try:
        from apps.employees.utils import get_holidays_and_short_days
        _, short_days_set = get_holidays_and_short_days(y, m)
    except Exception:
        short_days_set = set()
    day_types = []
    for d in days:
        dd = datetime.date(y, m, d)
        if dd in holidays:
            day_types.append('holiday')
        elif dd.weekday() >= 5:
            day_types.append('weekend')
        elif dd in short_days_set:
            day_types.append('short')
        else:
            day_types.append('work')

    days_with_types = list(zip(days, day_types))
    CODES = [
        ('Я','Явка'), ('ОТ','Отпуск'), ('ДО','Доп.отпуск'),
        ('УЧ','Учебный отпуск'), ('ОЖ','Отпуск по БиР'),
        ('Б','Больничный'), ('К','Командировка'), ('НН','Неявка'),
        ('П','Праздник'), ('В','Выходной'), ('Я½','Неполный'),
        ('РВ','Работа в выходной'), ('Я/С','Сверхурочные'),
    ]

    month_names = ["Январь","Февраль","Март","Апрель","Май","Июнь",
                   "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

    import json as _json
    from django.utils.safestring import mark_safe
    day_types_json = mark_safe(_json.dumps(day_types, ensure_ascii=False))
    return render(request, "dashboard/timesheet.html", {
        "employees": employees,
        "days": days,
        "day_types": day_types,
        "day_types_json": day_types_json,
        "days_with_types": days_with_types,
        "rec_map_json": _rec_map_to_json(rec_map),
        "year": y,
        "month": m,
        "month_name": month_names[m-1],
        "codes": CODES,
    })


def _rec_map_to_json(rec_map):
    import json
    d = {}
    for (emp_id, day), rec in rec_map.items():
        d[f"{emp_id}_{day}"] = {"code": rec.code, "hours": rec.hours}
    return json.dumps(d, ensure_ascii=False)


def _get_ru_holidays_dashboard(year):
    """Обёртка _get_ru_holidays для использования в views."""
    from apps.documents.services import _get_ru_holidays
    return _get_ru_holidays(year)


@login_required
@subscription_required
@require_role("hr")
def timesheet_save(request):
    """Сохраняет отметки табеля из POST (JSON body)."""
    import json
    from apps.employees.models import TimeRecord
    from datetime import date as dt_date
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({"error": "POST required"}, status=405)
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        from django.http import JsonResponse
        return JsonResponse({"error": "no company"}, status=403)
    try:
        data = json.loads(request.body)
    except Exception:
        from django.http import JsonResponse
        return JsonResponse({"error": "invalid json"}, status=400)

    # data = {"records": [{"employee_id": 1, "date": "2026-04-15", "code": "ОТ", "hours": 8}, ...]}
    employee_ids = set(Employee.objects.filter(company=member.company).values_list('id', flat=True))
    saved = 0
    for rec in data.get("records", []):
        emp_id = rec.get("employee_id")
        if emp_id not in employee_ids:
            continue
        try:
            d = dt_date.fromisoformat(rec["date"])
        except Exception:
            continue
        # Если clear=true — удалить запись (ячейка очищена)
        if rec.get("clear") or rec.get("code") == "":
            TimeRecord.objects.filter(employee_id=emp_id, date=d).delete()
            saved += 1
            continue
        code = rec.get("code", "Я")[:3]
        hours = int(rec.get("hours", 8))
        TimeRecord.objects.update_or_create(
            employee_id=emp_id, date=d,
            defaults={"code": code, "hours": hours}
        )
        saved += 1
    from django.http import JsonResponse
    return JsonResponse({"saved": saved})


# ─── СБРОС ПАРОЛЯ ──────────────────────────────────────────────────────────

def forgot_password_view(request):
    """Запрос ссылки для сброса пароля."""
    if request.user.is_authenticated:
        return redirect("dashboard:employees")

    if request.method == "POST":
        import uuid, json as _json, datetime
        import redis as _redis
        from django.conf import settings
        from django.utils import timezone

        email = request.POST.get("email", "").strip().lower()
        if not email:
            return render(request, "dashboard/forgot_password.html",
                          {"error": "Введите email"})

        # Не раскрываем, есть ли такой пользователь (защита от перебора)
        user = User.objects.filter(email=email).first()
        if user:
            token = str(uuid.uuid4())
            r = _redis.from_url(getattr(settings, "REDIS_RELAY_URL", "redis://redis:6379/2"))
            r.setex(f"password_reset:{token}", 3600, _json.dumps({
                "user_id": user.id,
                "email": email,
            }))
            reset_url = request.build_absolute_uri(f"/dashboard/reset-password/{token}/")
            from apps.accounts.tasks import send_password_reset_email
            send_password_reset_email.delay(email, reset_url)

        return render(request, "dashboard/forgot_password.html", {
            "success": f"Если аккаунт с адресом {email} существует, письмо будет отправлено в течение нескольких минут."
        })

    return render(request, "dashboard/forgot_password.html")


def reset_password_view(request, token):
    """Смена пароля по ссылке из письма."""
    from django.conf import settings
    import redis as _redis, json as _json
    from django.contrib.auth.hashers import make_password

    r = _redis.from_url(getattr(settings, "REDIS_RELAY_URL", "redis://redis:6379/2"))
    data_raw = r.get(f"password_reset:{token}")

    if not data_raw:
        return render(request, "dashboard/reset_password.html",
                      {"valid_token": False, "token": token})

    data = _json.loads(data_raw)

    if request.method == "POST":
        post_token = request.POST.get("token", "")
        password   = request.POST.get("password", "")
        password2  = request.POST.get("password2", "")

        if len(password) < 8:
            return render(request, "dashboard/reset_password.html",
                          {"valid_token": True, "token": token,
                           "error": "Пароль должен быть не менее 8 символов"})
        if password != password2:
            return render(request, "dashboard/reset_password.html",
                          {"valid_token": True, "token": token,
                           "error": "Пароли не совпадают"})

        try:
            user = User.objects.get(id=data["user_id"])
            user.set_password(password)
            user.save()
            r.delete(f"password_reset:{token}")
            # Автологин после смены пароля
            from django.contrib.auth import login as auth_login
            auth_login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            return redirect("dashboard:employees")
        except User.DoesNotExist:
            return render(request, "dashboard/reset_password.html",
                          {"valid_token": False, "token": token})

    return render(request, "dashboard/reset_password.html",
                  {"valid_token": True, "token": token})


@login_required
def change_password_view(request):
    """Смена пароля для авторизованного пользователя."""
    from django.contrib.auth import update_session_auth_hash

    if request.method == "POST":
        old_password = request.POST.get("old_password", "")
        password     = request.POST.get("password", "")
        password2    = request.POST.get("password2", "")

        if not request.user.check_password(old_password):
            return render(request, "dashboard/change_password.html",
                          {"error": "Текущий пароль введён неверно"})
        if len(password) < 8:
            return render(request, "dashboard/change_password.html",
                          {"error": "Новый пароль должен быть не менее 8 символов"})
        if password != password2:
            return render(request, "dashboard/change_password.html",
                          {"error": "Пароли не совпадают"})

        request.user.set_password(password)
        request.user.save()
        # Обновляем сессию чтобы не разлогинило
        update_session_auth_hash(request, request.user)
        return render(request, "dashboard/change_password.html",
                      {"success": "Пароль успешно изменён!"})

    return render(request, "dashboard/change_password.html")


# ─── ФОРМЫ И ДОКУМЕНТЫ ────────────────────────────────────────────────────

@login_required
@subscription_required
def forms_list(request):
    """Журнал всех документов компании с фильтром по типу"""
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect("dashboard:employees")
    company = member.company
    from apps.documents.models import Document
    doc_type = request.GET.get('type', '')
    documents = Document.objects.filter(company=company).select_related('employee').order_by('-created_at')
    if doc_type:
        documents = documents.filter(doc_type=doc_type)
    employees = Employee.objects.filter(company=company, status='active').order_by('last_name')
    return render(request, 'dashboard/forms_list.html', {
        'documents': documents,
        'employees': employees,
        'active_type': doc_type,
    })



@require_POST
@login_required
@require_role("admin")
def delete_document(request, doc_id):
    """Удаление документа из журнала"""
    from django.http import JsonResponse
    from apps.documents.models import Document
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return JsonResponse({'success': False, 'error': 'Компания не найдена'}, status=403)
    try:
        doc = Document.objects.get(id=doc_id, company=member.company)
        doc.delete()
        return JsonResponse({'success': True})
    except Document.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Документ не найден'}, status=404)


def _next_doc_number(company, doc_type):
    """Генерирует следующий номер документа по типу для компании."""
    from apps.documents.models import Document
    PREFIX = {
        "vacation":     "О",
        "gph_contract": "ГПХ",
        "gph_act":      "АКТ",
        "reference":    "С",
        "salary_change":"З",
        "transfer":     "ПР",
        "dismissal":    "У",
        "bonus":        "П",
        "disciplinary": "ДВ",
    }
    prefix = PREFIX.get(doc_type, "Д")
    existing = Document.objects.filter(company=company, doc_type=doc_type).values_list("number", flat=True)
    max_n = 0
    for num in existing:
        if num:
            import re
            m = re.search(r"(\d+)", str(num))
            if m:
                max_n = max(max_n, int(m.group(1)))
    return f"{prefix}-{max_n + 1}"


@login_required
@subscription_required
def form_editor(request, doc_type):
    """Редактор формы (новый или существующий документ)"""
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect("dashboard:employees")
    company = member.company
    from apps.documents.models import Document
    doc_id = request.GET.get('doc_id')
    document = None
    if doc_id:
        document = get_object_or_404(Document, id=doc_id, company=company)
    employees = Employee.objects.filter(company=company, status='active').order_by('last_name')

    FORM_TITLES = {
        'vacation': 'Приказ об отпуске (Т-6)',
        'gph_contract': 'Договор ГПХ',
        'gph_act': 'Акт выполненных работ',
        'reference': 'Справка с места работы',
        'salary_change': 'Изменение оклада',
        'transfer': 'Приказ о переводе (Т-5)',
        'dismissal': 'Приказ об увольнении (Т-8)',
        'bonus': 'Приказ о премии',
        'disciplinary': 'Приказ о дисциплинарном взыскании',
    }
    if doc_type not in FORM_TITLES:
        from django.http import Http404
        raise Http404

    next_number = _next_doc_number(company, doc_type) if not document else (document.number or _next_doc_number(company, doc_type))
    return render(request, 'dashboard/form_editor.html', {
        'doc_type': doc_type,
        'form_title': FORM_TITLES[doc_type],
        'document': document,
        'employees': employees,
        'extra_data': document.extra_data if document else {},
        'next_number': next_number,
        'company': company,
    })


@login_required
@subscription_required
@require_POST
@require_role("hr")
def form_save(request, doc_type):
    """Сохранение документа"""
    import json as _json_save
    from apps.documents.models import Document
    from django.http import JsonResponse
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return JsonResponse({'error': 'no company'}, status=403)
    company = member.company
    doc_id = request.POST.get('doc_id')
    employee_id = request.POST.get('employee_id')

    employee = None
    if employee_id:
        employee = get_object_or_404(Employee, id=employee_id, company=company)

    # Собираем extra_data из POST
    skip_keys = {'csrfmiddlewaretoken', 'doc_id', 'employee_id', 'doc_number', 'doc_date'}
    extra_data = {}
    for key, value in request.POST.items():
        if key not in skip_keys and value:
            extra_data[key] = value

    doc_number = request.POST.get('doc_number', '')
    doc_date_str = request.POST.get('doc_date', '')

    doc_date = date.today()
    if doc_date_str:
        parsed = _parse_date_flexible(doc_date_str)
        if parsed:
            doc_date = parsed

    FORM_TITLES = {
        'vacation': 'vacation',
        'gph_contract': 'gph_contract',
        'gph_act': 'gph_act',
        'reference': 'reference',
        'salary_change': 'salary_change',
        'transfer': 'transfer',
        'dismissal': 'dismissal',
        'bonus': 'bonus',
        'disciplinary': 'disciplinary',
    }
    if doc_type not in FORM_TITLES:
        return JsonResponse({'error': 'invalid doc_type'}, status=400)

    if doc_id:
        document = get_object_or_404(Document, id=doc_id, company=company)
        document.employee = employee
        document.number = doc_number
        document.date = doc_date
        document.extra_data = extra_data
        document.save()
    else:
        document = Document.objects.create(
            company=company,
            employee=employee,
            doc_type=doc_type,
            number=doc_number,
            date=doc_date,
            extra_data=extra_data,
        )

    # При сохранении salary_change — записываем старый оклад в историю и обновляем employee.salary
    if doc_type == 'salary_change' and employee:
        new_salary_raw = extra_data.get('new_salary')
        if new_salary_raw:
            try:
                new_salary_val = Decimal(new_salary_raw)
            except (InvalidOperation, ValueError):
                new_salary_val = None
            if new_salary_val and employee.salary != new_salary_val:
                from apps.employees.models import SalaryHistory
                old_salary = employee.salary  # сохраняем СТАРЫЙ оклад ДО обновления
                effective = _parse_date_flexible(extra_data.get('change_date')) or date.today()
                if old_salary:
                    SalaryHistory.objects.create(
                        employee=employee,
                        salary=old_salary,
                        effective_date=effective,
                    )
                # Сохраняем старый оклад в extra_data документа для PDF
                extra_data['old_salary'] = str(old_salary) if old_salary else '0'
                document.extra_data = extra_data
                document.save()
                # Только ПОТОМ обновляем оклад сотрудника
                employee.salary = new_salary_val
                employee.save()

    # При сохранении dismissal — обновляем fire_date и статус сотрудника
    if doc_type == 'dismissal' and employee:
        dismissal_date_raw = extra_data.get('dismissal_date')
        fire_date = _parse_date_flexible(dismissal_date_raw) or date.today()
        employee.fire_date = fire_date
        employee.status = 'fired'
        employee.save()

    # При сохранении transfer — обновляем должность и оклад сотрудника
    if doc_type == 'transfer' and employee:
        new_pos = extra_data.get('new_position')
        if new_pos:
            # Сохраняем старые данные в extra_data для PDF
            extra_data['old_position'] = employee.position or ''
            extra_data['old_salary'] = str(employee.salary) if employee.salary else ''
            document.extra_data = extra_data
            document.save()
            employee.position = new_pos
        new_sal = extra_data.get('new_salary')
        if new_sal:
            try:
                employee.salary = Decimal(new_sal)
            except (InvalidOperation, ValueError):
                pass
        employee.save()

    return JsonResponse({'success': True, 'doc_id': document.id, 'doc_type': doc_type})


@login_required
def employee_data_api(request, employee_id):
    """JSON данные сотрудника для автозаполнения"""
    from django.http import JsonResponse
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return JsonResponse({'error': 'no company'}, status=403)
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)
    # Последняя запись из истории окладов
    from apps.employees.models import SalaryHistory
    last_hist = employee.salary_history.order_by('-effective_date', '-created_at').first()
    data = {
        'id': employee.id,
        'full_name': employee.full_name,
        'first_name': employee.first_name,
        'last_name': employee.last_name,
        'middle_name': employee.middle_name or '',
        'position': employee.position or '',
        'salary': str(employee.salary) if employee.salary else '',
        'hire_date': employee.hire_date.strftime('%Y-%m-%d') if employee.hire_date else '',
        'inn': employee.inn or '',
        'snils': employee.snils or '',
        'phone': employee.phone or '',
        'email': employee.email or '',
        'passport_series': employee.passport_series or '',
        'passport_number': employee.passport_number or '',
        'previous_salary': str(last_hist.salary) if last_hist else '',
        'previous_salary_date': last_hist.effective_date.strftime('%d.%m.%Y') if last_hist else '',
    }
    return JsonResponse(data)


import hashlib as _hashlib
import time as _time

_CHAT_REDIS = 'redis://redis:6379/3'   # db=3 — чат (отдельно от celery и relay)
_BOT_TOKEN  = '7718001813:AAH4KBXZId8CurJdxpmno9jCJr5Bgcx01mM'
_ADMIN_CHAT = 1113292310


def _get_chat_session(request):
    """Уникальный 8-символьный ID сессии по email пользователя."""
    uid = request.user.email if request.user.is_authenticated else (
        request.session.session_key or 'anon'
    )
    return _hashlib.md5(uid.encode()).hexdigest()[:8]


def _chat_redis():
    import redis as _r
    return _r.from_url(_CHAT_REDIS)


def _save_msg(session_id, role, text, email=''):
    """Сохраняет сообщение в историю и обновляет индекс сессий."""
    import json as _j
    r = _chat_redis()
    msg = _j.dumps({'role': role, 'text': text, 'ts': int(_time.time())},
                   ensure_ascii=False)
    hist_key = f'chat_hist:{session_id}'
    r.rpush(hist_key, msg)
    r.expire(hist_key, 86400 * 7)   # храним 7 дней

    # Индекс всех сессий: hash  session_id -> {email, last_msg, ts}
    meta = _j.dumps({'email': email, 'last': text[:80], 'ts': int(_time.time())},
                    ensure_ascii=False)
    r.hset('chat_sessions', session_id, meta)


@require_POST
def chat_support(request):
    """Принимает сообщение клиента, сохраняет в историю, шлёт в Telegram."""
    from django.http import JsonResponse
    import json as _json, requests as _req

    try:
        data = _json.loads(request.body)
        text = data.get('text', '').strip()
    except Exception:
        text = request.POST.get('text', '').strip()

    if not text:
        return JsonResponse({'ok': False, 'error': 'empty'})

    user_email = request.user.email if request.user.is_authenticated else 'аноним'
    session_id = _get_chat_session(request)

    # Сохраняем сообщение клиента в историю
    _save_msg(session_id, 'user', text, email=user_email)

    tg_text = (
        f'\U0001f4ac Чат поддержки [{session_id}]\n'
        f'\U0001f464 {user_email}\n'
        f'\U0001f4dd {text}\n\n'
        f'<i>Reply на это сообщение чтобы ответить клиенту</i>'
    )

    try:
        resp = _req.post(
            f'https://api.telegram.org/bot{_BOT_TOKEN}/sendMessage',
            json={'chat_id': _ADMIN_CHAT, 'text': tg_text, 'parse_mode': 'HTML'},
            timeout=10,
        )
        result = resp.json()
        return JsonResponse({'ok': result.get('ok', False), 'session': session_id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


def chat_history(request):
    """Возвращает историю переписки для сессии (вызывается при загрузке страницы)."""
    from django.http import JsonResponse
    import json as _json

    session_id = request.GET.get('session', '')
    if not session_id or len(session_id) != 8:
        # Вычисляем session_id из текущего пользователя
        session_id = _get_chat_session(request)

    r = _chat_redis()
    hist_key = f'chat_hist:{session_id}'
    raw_list = r.lrange(hist_key, 0, -1)

    messages = []
    for raw in raw_list:
        try:
            messages.append(_json.loads(raw))
        except Exception:
            pass

    return JsonResponse({'session': session_id, 'messages': messages})


@csrf_exempt
def chat_webhook(request):
    """Webhook от Telegram — ответы оператора и команды бота."""
    from django.http import JsonResponse
    import json as _json, re as _re, requests as _req

    if request.method != 'POST':
        return JsonResponse({'ok': True})

    try:
        data = _json.loads(request.body)
    except Exception:
        return JsonResponse({'ok': False})

    msg = data.get('message', {})
    if not msg:
        return JsonResponse({'ok': True})

    from_chat = msg.get('chat', {}).get('id')
    text = msg.get('text', '').strip()

    # ── Команда /чаты ──────────────────────────────────────────────────────
    if text in ('/чаты', '/chats', '/чаты@kadrovik_leads_bot'):
        r = _chat_redis()
        sessions = r.hgetall('chat_sessions')

        if not sessions:
            reply = '\U0001f4c2 Активных чатов нет.'
        else:
            import json as _j2
            lines = ['\U0001f4cb <b>Активные сессии чата:</b>\n']
            for sid, meta_raw in sessions.items():
                sid = sid.decode() if isinstance(sid, bytes) else sid
                try:
                    meta = _j2.loads(meta_raw)
                except Exception:
                    continue
                import datetime
                ts = datetime.datetime.fromtimestamp(meta['ts']).strftime('%d.%m %H:%M')
                lines.append(
                    f'\U0001f464 <b>{meta.get("email","?")}</b> [{sid}]\n'
                    f'   \U0001f4dd {meta.get("last","...")}\n'
                    f'   \U0001f552 {ts}\n'
                )
            reply = '\n'.join(lines)

        _req.post(
            f'https://api.telegram.org/bot{_BOT_TOKEN}/sendMessage',
            json={'chat_id': from_chat, 'text': reply, 'parse_mode': 'HTML'},
            timeout=10,
        )
        return JsonResponse({'ok': True})

    # ── Ответ оператора клиенту ────────────────────────────────────────────
    session_id = None

    # Способ 1: reply на сообщение бота
    reply_to = msg.get('reply_to_message', {})
    if reply_to:
        orig = reply_to.get('text', '')
        m = _re.search(r'\[([a-f0-9]{8})\]', orig)
        if m:
            session_id = m.group(1)

    # Способ 2: оператор пишет [session_id] текст
    if not session_id:
        m = _re.match(r'\[([a-f0-9]{8})\]\s*(.*)', text, _re.DOTALL)
        if m:
            session_id = m.group(1)
            text = m.group(2).strip()

    if not session_id or not text:
        return JsonResponse({'ok': True})

    # Сохраняем ответ в историю
    _save_msg(session_id, 'bot', text)

    # Кладём в очередь для poll
    r = _chat_redis()
    import json as _j3
    r.rpush(f'chat_replies:{session_id}',
            _j3.dumps({'text': text, 'ts': int(_time.time())}, ensure_ascii=False))
    r.expire(f'chat_replies:{session_id}', 3600)

    return JsonResponse({'ok': True})


def chat_poll(request):
    """Клиент опрашивает новые сообщения от оператора (только новые, не история)."""
    from django.http import JsonResponse
    import json as _json

    session_id = request.GET.get('session', '')
    if not session_id or len(session_id) != 8:
        return JsonResponse({'messages': []})

    r = _chat_redis()
    key = f'chat_replies:{session_id}'
    messages = []
    while True:
        raw = r.lpop(key)
        if not raw:
            break
        try:
            messages.append(_json.loads(raw))
        except Exception:
            pass

    return JsonResponse({'messages': messages})


@require_POST
@login_required
@require_role("admin")
def delete_document(request, doc_id):
    from django.http import JsonResponse
    from apps.documents.models import Document
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return JsonResponse({"success": False, "error": "Нет доступа"}, status=403)
    try:
        doc = Document.objects.get(id=doc_id, company=member.company)
        doc.delete()
        return JsonResponse({"success": True})
    except Document.DoesNotExist:
        return JsonResponse({"success": False, "error": "Документ не найден"}, status=404)


# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

def _require_plan(request, min_plan):
    """Проверяет тариф пользователя. Возвращает True если доступ разрешён."""
    PLAN_RANK = {'trial': 0, 'start': 1, 'business': 2, 'pro': 3}
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return False
    sub = getattr(member.company, 'subscription', None)
    if not sub or not sub.is_active:
        return False
    return PLAN_RANK.get(sub.plan, 0) >= PLAN_RANK.get(min_plan, 2)


@login_required
@subscription_required
def export_employees_excel(request):
    """Экспорт списка сотрудников в Excel. Тариф Бизнес+."""
    from django.http import HttpResponse, JsonResponse
    if not _require_plan(request, 'business'):
        return JsonResponse({'error': 'Функция доступна на тарифах Бизнес и Корпоратив'}, status=403)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    member = CompanyMember.objects.filter(user=request.user).first()
    company = member.company
    employees = Employee.objects.filter(company=company).select_related('department').order_by('last_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Сотрудники'

    header_fill = PatternFill('solid', fgColor='1E40AF')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    cell_font   = Font(size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    ws.merge_cells('A1:K1')
    ws['A1'] = f'{company.name} — Список сотрудников'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ['#', 'Фамилия', 'Имя', 'Отчество', 'Должность', 'Отдел', 'Таб. номер', 'Дата приёма', 'Оклад', 'Телефон', 'Статус']
    col_widths = [5, 18, 14, 14, 22, 18, 13, 14, 12, 16, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22

    STATUS_LABELS = {'active': 'Работает', 'fired': 'Уволен', 'vacation': 'В отпуске'}

    for row_idx, emp in enumerate(employees, 3):
        row = [
            row_idx - 2,
            emp.last_name or '',
            emp.first_name or '',
            emp.middle_name or '',
            emp.position or '',
            emp.department.name if emp.department else '',
            emp.personnel_number or '',
            emp.hire_date.strftime('%d.%m.%Y') if emp.hire_date else '',
            f'{emp.salary:,.2f} ₽' if emp.salary else '',
            emp.phone or '',
            STATUS_LABELS.get(emp.status, emp.status),
        ]
        alt_fill = PatternFill('solid', fgColor='EFF6FF') if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = border
            cell.alignment = center if col_idx in (1, 7, 8, 9, 11) else left
            if alt_fill:
                cell.fill = alt_fill

    total_row = len(employees) + 3
    ws.cell(row=total_row, column=1, value='ИТОГО:')
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=2, value=f'{employees.count()} чел.')
    ws.cell(row=total_row, column=2).font = Font(bold=True, size=10)

    ws.freeze_panes = 'A3'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date as _d
    filename = f'employees_{company.name}_{_d.today().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@subscription_required
def export_timesheet_excel(request):
    """Экспорт табеля в Excel. Тариф Бизнес+."""
    import calendar
    import datetime
    from django.http import HttpResponse, JsonResponse
    from apps.employees.models import TimeRecord
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not _require_plan(request, 'business'):
        return JsonResponse({'error': 'Функция доступна на тарифах Бизнес и Корпоратив'}, status=403)

    member = CompanyMember.objects.filter(user=request.user).first()
    company = member.company

    try:
        y = int(request.GET.get('year', datetime.date.today().year))
        m = int(request.GET.get('month', datetime.date.today().month))
    except (ValueError, TypeError):
        y, m = datetime.date.today().year, datetime.date.today().month

    month_names = ['Январь','Февраль','Март','Апрель','Май','Июнь',
                   'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
    days_in_month = calendar.monthrange(y, m)[1]
    days = list(range(1, days_in_month + 1))

    employees = Employee.objects.filter(company=company).select_related('department').order_by('last_name')
    start = datetime.date(y, m, 1)
    end = datetime.date(y, m, days_in_month)
    records = TimeRecord.objects.filter(employee__in=employees, date__gte=start, date__lte=end)
    rec_map = {(r.employee_id, r.date.day): r for r in records}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Табель {month_names[m-1]} {y}'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    subhead_fill = PatternFill('solid', fgColor='2563EB')
    weekend_fill = PatternFill('solid', fgColor='FEE2E2')
    ya_fill      = PatternFill('solid', fgColor='D1FAE5')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    hfont  = Font(bold=True, color='FFFFFF', size=9)
    sfont  = Font(size=9)

    total_cols = 4 + days_in_month + 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1).value = f'{company.name} — Табель (Т-13) — {month_names[m-1]} {y}'
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = 26

    fixed_headers = ['#', 'ФИО', 'Должность', 'Отдел']
    col_w = [4, 22, 18, 14]
    for ci, (h, w) in enumerate(zip(fixed_headers, col_w), 1):
        c = ws.cell(2, ci, h)
        c.fill = header_fill; c.font = hfont; c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    holidays = _get_ru_holidays_dashboard(y)
    try:
        from apps.employees.utils import get_holidays_and_short_days
        _, short_days = get_holidays_and_short_days(y, m)
    except Exception:
        short_days = set()
    for di, d in enumerate(days, 5):
        dd = datetime.date(y, m, d)
        is_wknd = dd.weekday() >= 5 or dd in holidays
        c = ws.cell(2, di, d)
        c.fill = weekend_fill if is_wknd else subhead_fill
        c.font = Font(bold=True, color='000000' if is_wknd else 'FFFFFF', size=8)
        c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(di)].width = 4.5

    ws.cell(2, 5 + days_in_month, 'Дней').fill = subhead_fill
    ws.cell(2, 5 + days_in_month).font = hfont
    ws.cell(2, 5 + days_in_month).alignment = center
    ws.cell(2, 5 + days_in_month).border = border
    ws.column_dimensions[get_column_letter(5 + days_in_month)].width = 7

    ws.cell(2, 6 + days_in_month, 'Часов').fill = subhead_fill
    ws.cell(2, 6 + days_in_month).font = hfont
    ws.cell(2, 6 + days_in_month).alignment = center
    ws.cell(2, 6 + days_in_month).border = border
    ws.column_dimensions[get_column_letter(6 + days_in_month)].width = 7

    ws.row_dimensions[2].height = 20

    for ri, emp in enumerate(employees, 3):
        ws.cell(ri, 1, ri - 2).font = sfont; ws.cell(ri, 1).alignment = center; ws.cell(ri, 1).border = border
        ws.cell(ri, 2, emp.full_name).font = sfont; ws.cell(ri, 2).border = border
        ws.cell(ri, 3, emp.position or '').font = sfont; ws.cell(ri, 3).border = border
        ws.cell(ri, 4, emp.department.name if emp.department else '').font = sfont; ws.cell(ri, 4).border = border

        total_days = 0
        total_hours = 0
        for di, d in enumerate(days, 5):
            rec = rec_map.get((emp.id, d))
            dd = datetime.date(y, m, d)
            is_wknd = dd.weekday() >= 5 or dd in holidays
            is_short = dd in short_days
            code = rec.code if rec else ('В' if is_wknd else '')
            hours = rec.hours if rec else 0
            c = ws.cell(ri, di, code)
            c.font = Font(size=8)
            c.alignment = center
            c.border = border
            if is_wknd and not rec:
                c.fill = weekend_fill
            elif code == 'Я':
                c.fill = ya_fill
                total_days += 1
                total_hours += hours or (7 if is_short else 8)
            elif code and code not in ('В', 'П', ''):
                total_days += 1
                total_hours += hours or 0

        ws.cell(ri, 5 + days_in_month, total_days).font = Font(bold=True, size=9)
        ws.cell(ri, 5 + days_in_month).alignment = center
        ws.cell(ri, 5 + days_in_month).border = border
        ws.cell(ri, 6 + days_in_month, total_hours).font = Font(bold=True, size=9)
        ws.cell(ri, 6 + days_in_month).alignment = center
        ws.cell(ri, 6 + days_in_month).border = border
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = 'E3'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'timesheet_{y}_{m:02d}_{company.name}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ===== TEAM MANAGEMENT =====

from apps.companies.models import CompanyInvite
from django.contrib import messages as django_messages


@login_required
def team_list(request):
    """Страница управления командой."""
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    company = member.company
    members = CompanyMember.objects.filter(company=company).select_related('user').order_by('created_at')
    pending_invites = CompanyInvite.objects.filter(company=company, accepted=False).order_by('-created_at')
    context = {
        'members': members,
        'pending_invites': pending_invites,
        'current_member': member,
    }
    return render(request, 'dashboard/team.html', context)


@login_required
@require_role("admin")
def team_invite(request):
    """Отправить приглашение."""
    if request.method != 'POST':
        return redirect('dashboard:team_list')
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    company = member.company

    # Проверка тарифа
    from apps.billing.models import Subscription
    from apps.billing.services import PLANS
    subscription = Subscription.objects.filter(company=company).order_by('-started_at').first()
    plan = subscription.plan if subscription else 'trial'
    plan_features = PLANS.get(plan, PLANS['trial'])['features']
    if not plan_features.get('multi_user'):
        django_messages.error(request, 'Функция доступна с тарифа Бизнес.')
        return redirect('dashboard:team_list')

    email = request.POST.get('email', '').strip().lower()
    role = request.POST.get('role', 'hr')

    if not email:
        django_messages.error(request, 'Введите email.')
        return redirect('dashboard:team_list')

    # Проверка: пользователь уже в команде?
    existing_user = User.objects.filter(email=email).first()
    if existing_user and CompanyMember.objects.filter(company=company, user=existing_user).exists():
        django_messages.warning(request, 'Этот пользователь уже в вашей команде.')
        return redirect('dashboard:team_list')

    # Создать или обновить приглашение
    from django.utils import timezone
    from datetime import timedelta as td
    invite, created = CompanyInvite.objects.update_or_create(
        company=company,
        email=email,
        defaults={
            'invited_by': request.user,
            'role': role,
            'accepted': False,
            'expires_at': timezone.now() + td(days=7),
        }
    )

    # Отправить email
    invite_url = request.build_absolute_uri(f'/dashboard/invite/{invite.token}/')
    from django.core.mail import send_mail
    from django.conf import settings as django_settings
    subject = f'Приглашение в {company.name} — Кадровый автопилот'
    plain = f'Вас приглашают присоединиться к компании {company.name}.\nПерейдите по ссылке: {invite_url}'
    html = f"""
    <html><body style="font-family: Arial, sans-serif; color: #333;">
    <div style="max-width:600px;margin:0 auto;padding:20px;">
    <h2 style="color:#1e3a5f;">Приглашение в команду</h2>
    <p>Вас приглашают присоединиться к компании <strong>{company.name}</strong> в системе «Кадровый автопилот».</p>
    <p style="margin:30px 0;">
      <a href="{invite_url}" style="background:#1e3a5f;color:#fff;padding:12px 28px;border-radius:6px;text-decoration:none;font-size:16px;">Принять приглашение</a>
    </p>
    <p style="color:#888;font-size:13px;">Ссылка действует 7 дней. Если вы не ожидали этого письма — просто проигнорируйте его.</p>
    <hr style="border:none;border-top:1px solid #eee;margin-top:30px;">
    <p style="color:#aaa;font-size:12px;">Кадровый автопилот | kadrovik-auto.ru</p>
    </div></body></html>
    """
    try:
        send_mail(subject, plain, django_settings.DEFAULT_FROM_EMAIL, [email], html_message=html, fail_silently=False)
        django_messages.success(request, f'Приглашение отправлено на {email}.')
    except Exception as e:
        django_messages.error(request, f'Ошибка отправки email: {e}')

    return redirect('dashboard:team_list')


def invite_accept(request, token):
    """Принятие приглашения (публичный URL)."""
    from django.utils import timezone
    invite = get_object_or_404(CompanyInvite, token=token, accepted=False)
    if invite.is_expired():
        return render(request, 'dashboard/invite_expired.html', {'invite': invite})

    if request.method == 'POST':
        if request.user.is_authenticated:
            user = request.user
        else:
            from django.contrib.auth import get_user_model
            UserModel = get_user_model()
            email = request.POST.get('email', '').strip().lower()
            password = request.POST.get('password', '')
            confirm = request.POST.get('confirm_password', '')
            is_new = request.POST.get('is_new') == '1'

            if is_new:
                if password != confirm:
                    return render(request, 'dashboard/invite_accept.html', {'invite': invite, 'error': 'Пароли не совпадают'})
                if len(password) < 6:
                    return render(request, 'dashboard/invite_accept.html', {'invite': invite, 'error': 'Пароль должен быть не менее 6 символов'})
                if UserModel.objects.filter(email=email).exists():
                    return render(request, 'dashboard/invite_accept.html', {'invite': invite, 'error': 'Пользователь с таким email уже существует. Войдите в аккаунт.'})
                user = UserModel.objects.create_user(email=email, password=password, username=email)
            else:
                user = authenticate(request, email=email, password=password)
                if not user:
                    user = authenticate(request, username=email, password=password)
                if not user:
                    return render(request, 'dashboard/invite_accept.html', {'invite': invite, 'error': 'Неверный email или пароль'})
            login(request, user)

        # Добавить в компанию
        CompanyMember.objects.get_or_create(
            company=invite.company,
            user=request.user,
            defaults={'role': invite.role}
        )
        invite.accepted = True
        invite.save()
        django_messages.success(request, f'Вы присоединились к компании {invite.company.name}!')
        return redirect('dashboard:home')

    # GET — показать форму
    already_member = request.user.is_authenticated and CompanyMember.objects.filter(company=invite.company, user=request.user).exists()
    return render(request, 'dashboard/invite_accept.html', {'invite': invite, 'already_member': already_member})


@login_required
@require_role("admin")
def team_member_remove(request, member_id):
    """Удалить участника команды."""
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    target = get_object_or_404(CompanyMember, id=member_id, company=member.company)
    if target.role == 'owner':
        django_messages.error(request, 'Нельзя удалить владельца компании.')
        return redirect('dashboard:team_list')
    if target.user == request.user:
        django_messages.error(request, 'Нельзя удалить себя.')
        return redirect('dashboard:team_list')
    target.delete()
    django_messages.success(request, 'Участник удалён из команды.')
    return redirect('dashboard:team_list')


@login_required
@require_role("admin")
def team_invite_cancel(request, invite_id):
    """Отменить приглашение."""
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    invite = get_object_or_404(CompanyInvite, id=invite_id, company=member.company, accepted=False)
    invite.delete()
    django_messages.success(request, 'Приглашение отменено.')
    return redirect('dashboard:team_list')


@login_required
@require_role("admin")
def api_settings(request):
    """Страница управления API-токеном."""
    from apps.billing.services import PLANS
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    sub = getattr(member.company, 'subscription', None)
    plan = sub.plan if sub else 'trial'
    features = PLANS.get(plan, PLANS['trial'])['features']

    token = None
    if features.get('api'):
        from rest_framework.authtoken.models import Token
        token, _ = Token.objects.get_or_create(user=request.user)

    context = {
        'token': token,
        'has_api': features.get('api', False),
        'api_base_url': request.build_absolute_uri('/api/v1/'),
    }
    return render(request, 'dashboard/api_settings.html', context)


@login_required
@require_role("admin")
def api_token_regenerate(request):
    """Пересоздать API-токен."""
    if request.method != 'POST':
        return redirect('dashboard:api_settings')
    from rest_framework.authtoken.models import Token
    Token.objects.filter(user=request.user).delete()
    Token.objects.create(user=request.user)
    from django.contrib import messages
    messages.success(request, 'API-токен обновлён.')
    return redirect('dashboard:api_settings')


# ===== DOCUMENT TEMPLATES =====

@login_required
def document_templates(request):
    """Страница управления кастомными шаблонами документов."""
    from apps.billing.models import Subscription
    from apps.billing.services import PLANS
    from apps.documents.models import DocumentTemplate, DOC_TYPES

    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    company = member.company

    sub = getattr(company, 'subscription', None)
    plan = sub.plan if sub else 'trial'
    features = PLANS.get(plan, PLANS['trial'])['features']

    templates_by_type = {t.doc_type: t for t in DocumentTemplate.objects.filter(company=company)}

    context = {
        'has_custom_templates': features.get('custom_templates', False),
        'doc_types': DOC_TYPES,
        'templates_by_type': templates_by_type,
    }
    return render(request, 'dashboard/document_templates.html', context)


@login_required
def document_template_upload(request, doc_type):
    """Загрузить кастомный шаблон для типа документа."""
    if request.method != 'POST':
        return redirect('dashboard:document_templates')

    from apps.billing.models import Subscription
    from apps.billing.services import PLANS
    from apps.documents.models import DocumentTemplate

    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    company = member.company

    sub = getattr(company, 'subscription', None)
    plan = sub.plan if sub else 'trial'
    features = PLANS.get(plan, PLANS['trial'])['features']

    if not features.get('custom_templates'):
        django_messages.error(request, 'Кастомные шаблоны доступны только на тарифе Корпоратив.')
        return redirect('dashboard:document_templates')

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        django_messages.error(request, 'Файл не выбран.')
        return redirect('dashboard:document_templates')

    if not uploaded_file.name.endswith('.docx'):
        django_messages.error(request, 'Допускаются только файлы .docx.')
        return redirect('dashboard:document_templates')

    existing = DocumentTemplate.objects.filter(company=company, doc_type=doc_type).first()
    if existing:
        existing.delete()

    DocumentTemplate.objects.create(
        company=company,
        doc_type=doc_type,
        file=uploaded_file,
        name=uploaded_file.name,
    )
    django_messages.success(request, 'Шаблон загружен.')
    return redirect('dashboard:document_templates')


@login_required
def document_template_delete(request, doc_type):
    """Удалить кастомный шаблон (вернуться к стандартному)."""
    from apps.documents.models import DocumentTemplate
    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    template = get_object_or_404(DocumentTemplate, company=member.company, doc_type=doc_type)
    template.delete()
    django_messages.success(request, 'Шаблон удалён. Будет использоваться стандартный.')
    return redirect('dashboard:document_templates')


@login_required
def document_template_download(request, doc_type, employee_id):
    """Скачать документ на основе кастомного шаблона."""
    from apps.documents.models import DocumentTemplate
    from apps.documents.template_renderer import render_template_to_bytes, get_employee_context, get_company_context

    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')

    template = get_object_or_404(DocumentTemplate, company=member.company, doc_type=doc_type)
    employee = get_object_or_404(Employee, id=employee_id, company=member.company)

    context = {}
    context.update(get_employee_context(employee))
    context.update(get_company_context(member.company))

    for key in ['doc_number', 'doc_date', 'reason', 'note']:
        val = request.GET.get(key) or request.POST.get(key, '')
        if val:
            context[key] = val

    try:
        docx_bytes = render_template_to_bytes(template.file.path, context)
    except Exception as e:
        django_messages.error(request, f'Ошибка рендеринга шаблона: {e}')
        return redirect('dashboard:document_templates')

    filename = f"{doc_type}_{employee.last_name}.docx"
    response = HttpResponse(docx_bytes, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# ===== SFR EXPORT =====

@login_required
def sfr_export(request):
    from apps.billing.models import Subscription
    from apps.billing.services import PLANS
    from apps.employees.models import Employee

    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    company = member.company

    subscription = Subscription.objects.filter(company=company).order_by('-started_at').first()
    plan = subscription.plan if subscription else 'trial'
    features = PLANS.get(plan, PLANS['trial'])['features']

    context = {
        'has_sfr_export': features.get('sfr_export', False),
        'company': company,
    }

    if not features.get('sfr_export'):
        return render(request, 'dashboard/sfr_export.html', context)

    if request.method == 'POST':
        return _sfr_export_generate(request, company)

    # GET
    employees = Employee.objects.filter(company=company).order_by('last_name')

    period_start_str = request.GET.get('period_start', '')
    period_end_str = request.GET.get('period_end', '')

    import datetime as _dt
    today = date.today()
    try:
        period_start = _dt.datetime.strptime(period_start_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        period_start = date(today.year, today.month, 1)
    try:
        period_end = _dt.datetime.strptime(period_end_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        period_end = today

    events_preview = []
    for emp in employees:
        if emp.hire_date and period_start <= emp.hire_date <= period_end:
            events_preview.append({
                'employee': emp,
                'event_type': 'hire',
                'event_label': 'Приём',
                'event_date': emp.hire_date,
                'position': emp.position or '',
            })
        if emp.fire_date and period_start <= emp.fire_date <= period_end:
            events_preview.append({
                'employee': emp,
                'event_type': 'dismiss',
                'event_label': 'Увольнение',
                'event_date': emp.fire_date,
                'position': emp.position or '',
            })

    events_preview.sort(key=lambda x: x['event_date'])

    context.update({
        'events_preview': events_preview,
        'period_start': period_start.strftime('%Y-%m-%d'),
        'period_end': period_end.strftime('%Y-%m-%d'),
        'has_sfr_reg_number': bool(company.sfr_reg_number),
        'has_okved': bool(company.okved),
    })

    return render(request, 'dashboard/sfr_export.html', context)


def _sfr_export_generate(request, company):
    from apps.documents.sfr_generator import generate_efs1_xml
    from apps.employees.models import Employee
    import datetime as _dt

    period_start_str = request.POST.get('period_start', '')
    period_end_str = request.POST.get('period_end', '')

    today = date.today()
    try:
        period_start = _dt.datetime.strptime(period_start_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        period_start = date(today.year, today.month, 1)
    try:
        period_end = _dt.datetime.strptime(period_end_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        period_end = today

    employees = Employee.objects.filter(company=company)
    events = []
    for emp in employees:
        if emp.hire_date and period_start <= emp.hire_date <= period_end:
            events.append({
                'employee': emp,
                'event_type': 'hire',
                'event_date': emp.hire_date,
                'position': emp.position or '',
            })
        if emp.fire_date and period_start <= emp.fire_date <= period_end:
            events.append({
                'employee': emp,
                'event_type': 'dismiss',
                'event_date': emp.fire_date,
                'position': emp.position or '',
            })

    xml_bytes = generate_efs1_xml(company, events, period_start, period_end)

    filename = 'EFS1_' + (company.inn or 'noinn') + '_' + today.strftime('%Y%m%d') + '.xml'
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    return response


# ===== СОБЫТИЯ =====

@login_required
def events_list(request):
    """Страница ленты HR-событий компании."""
    from apps.events.models import HREvent
    from apps.employees.models import Employee
    from apps.vacations.models import Vacation
    from django.utils import timezone
    from datetime import timedelta

    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return redirect('dashboard:login')
    company = member.company
    employees = Employee.objects.filter(company=company, status='active')

    today = timezone.now().date()
    upcoming_days = 30

    events = []

    # 1. Дни рождения (в следующие 30 дней)
    for emp in employees:
        if not emp.birth_date:
            continue
        try:
            bday_this_year = emp.birth_date.replace(year=today.year)
        except ValueError:
            # 29 февраля
            bday_this_year = emp.birth_date.replace(year=today.year, day=28)
        if bday_this_year < today:
            try:
                bday_this_year = emp.birth_date.replace(year=today.year + 1)
            except ValueError:
                bday_this_year = emp.birth_date.replace(year=today.year + 1, day=28)
        days_until = (bday_this_year - today).days
        if 0 <= days_until <= upcoming_days:
            age = today.year - emp.birth_date.year + (1 if bday_this_year.year > today.year else 0)
            events.append({
                'type': 'birthday',
                'icon': '\U0001f382',
                'color': 'info',
                'title': f'\u0414\u0435\u043d\u044c \u0440\u043e\u0436\u0434\u0435\u043d\u0438\u044f \u2014 {emp.last_name} {emp.first_name}',
                'description': f'\u0418\u0441\u043f\u043e\u043b\u043d\u044f\u0435\u0442\u0441\u044f {age} \u043b\u0435\u0442',
                'date': bday_this_year,
                'days_until': days_until,
                'employee': emp,
                'urgency': 'today' if days_until == 0 else ('soon' if days_until <= 3 else 'upcoming'),
            })

    # 2. Испытательный срок (заканчивается в следующие 14 дней)
    for emp in employees:
        if not emp.probation_end_date:
            continue
        days_until = (emp.probation_end_date - today).days
        if 0 <= days_until <= 14:
            events.append({
                'type': 'probation',
                'icon': '\u23f0',
                'color': 'warning',
                'title': f'\u0418\u0441\u043f\u044b\u0442\u0430\u0442\u0435\u043b\u044c\u043d\u044b\u0439 \u0441\u0440\u043e\u043a \u2014 {emp.last_name} {emp.first_name}',
                'description': f'\u0417\u0430\u043a\u0430\u043d\u0447\u0438\u0432\u0430\u0435\u0442\u0441\u044f \u0447\u0435\u0440\u0435\u0437 {days_until} \u0434\u043d.',
                'date': emp.probation_end_date,
                'days_until': days_until,
                'employee': emp,
                'urgency': 'today' if days_until == 0 else ('soon' if days_until <= 3 else 'upcoming'),
            })

    # 3. Срочные договоры (заканчиваются в следующие 14 дней)
    for emp in employees:
        if emp.contract_type != 'fixed_term' or not emp.contract_end_date:
            continue
        days_until = (emp.contract_end_date - today).days
        if 0 <= days_until <= 14:
            events.append({
                'type': 'contract',
                'icon': '\U0001f4c4',
                'color': 'danger',
                'title': f'\u0421\u0440\u043e\u0447\u043d\u044b\u0439 \u0434\u043e\u0433\u043e\u0432\u043e\u0440 \u2014 {emp.last_name} {emp.first_name}',
                'description': f'\u0418\u0441\u0442\u0435\u043a\u0430\u0435\u0442 \u0447\u0435\u0440\u0435\u0437 {days_until} \u0434\u043d.',
                'date': emp.contract_end_date,
                'days_until': days_until,
                'employee': emp,
                'urgency': 'today' if days_until == 0 else ('soon' if days_until <= 3 else 'upcoming'),
            })

    # 4. Отпуска (начинаются в следующие 7 дней или текущие)
    try:
        vacations = Vacation.objects.filter(
            employee__company=company,
            start_date__lte=today + timedelta(days=7),
            end_date__gte=today
        ).select_related('employee')
        for vac in vacations:
            days_until = (vac.start_date - today).days
            events.append({
                'type': 'vacation',
                'icon': '\U0001f3d6\ufe0f',
                'color': 'success',
                'title': f'\u041e\u0442\u043f\u0443\u0441\u043a \u2014 {vac.employee.last_name} {vac.employee.first_name}',
                'description': vac.start_date.strftime('%d.%m') + ' — ' + vac.end_date.strftime('%d.%m.%Y'),
                'date': vac.start_date,
                'days_until': max(0, days_until),
                'employee': vac.employee,
                'urgency': 'today' if days_until <= 0 else ('soon' if days_until <= 2 else 'upcoming'),
            })
    except Exception:
        pass

    # Сортировать по дате
    events.sort(key=lambda x: (x['date'], x['days_until']))

    # Разбить на срочные и предстоящие
    urgent_events = [e for e in events if e['days_until'] <= 3]
    upcoming_events = [e for e in events if e['days_until'] > 3]

    context = {
        'events': events,
        'urgent_events': urgent_events,
        'upcoming_events': upcoming_events,
        'today': today,
    }
    return render(request, 'dashboard/events.html', context)


@login_required
def events_count_api(request):
    """JSON API — количество срочных событий (для колокольчика в header)."""
    from apps.employees.models import Employee
    from apps.vacations.models import Vacation
    from django.http import JsonResponse
    from django.utils import timezone
    from datetime import timedelta

    member = CompanyMember.objects.filter(user=request.user).first()
    if not member:
        return JsonResponse({'count': 0})
    company = member.company
    employees = Employee.objects.filter(company=company, status='active')
    today = timezone.now().date()
    count = 0

    for emp in employees:
        # Дни рождения <= 3 дней
        if emp.birth_date:
            try:
                bday = emp.birth_date.replace(year=today.year)
            except ValueError:
                bday = emp.birth_date.replace(year=today.year, day=28)
            if bday < today:
                try:
                    bday = emp.birth_date.replace(year=today.year + 1)
                except ValueError:
                    bday = emp.birth_date.replace(year=today.year + 1, day=28)
            if (bday - today).days <= 3:
                count += 1
        # Испытательный срок <= 7 дней
        if emp.probation_end_date:
            if 0 <= (emp.probation_end_date - today).days <= 7:
                count += 1
        # Срочный договор <= 7 дней
        if emp.contract_type == 'fixed_term' and emp.contract_end_date:
            if 0 <= (emp.contract_end_date - today).days <= 7:
                count += 1

    # Отпуска <= 3 дней
    try:
        vac_count = Vacation.objects.filter(
            employee__company=company,
            start_date__lte=today + timedelta(days=3),
            start_date__gte=today,
        ).count()
        count += vac_count
    except Exception:
        pass

    return JsonResponse({'count': count})
